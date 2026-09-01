"""Render the AI-capex forward-ROIC model to .xlsx from the data layer.

VALUES ONLY. ``model/calc.py`` is the single calculation engine; this script is a
renderer. It never writes an Excel formula that recomputes the model -- that would
put the model in two places and reintroduce the drift this architecture removed.
The only formulas that could legitimately live in the output are the ``Checks``
comparison formulas, and even those are rendered as values plus a literal
PASS/FAIL so the sheet is readable without a recalculating spreadsheet engine
(see ``_build_checks``).

What is carried forward from the frozen audit-of-record
(``ai_capex_forward_roic_analysis_v02.xlsx``):

* every cell note in ``data/cell_notes.csv`` (455 today), verbatim where the cell
  still exists; for cells that did not exist in v02 (new quarters) the note is
  produced by rewriting the *same ticker's* v02 note line-for-line with that
  quarter's own facts, so no URL or evidence sentence is ever invented;
* every hyperlink in ``data/hyperlinks.csv`` (240 today), rebuilt from the data
  layer's own URLs;
* the blue/yellow fill language -- blue ``FFEAF2F8`` = sourced filing fact,
  yellow ``FFFFF2CC`` = analyst assumption (``data/provenance.csv``);
* sheet order and names, number formats, column widths, row heights and the
  header/banner prose.

What changes versus v02:

1. ``Trajectory`` renders one column per quarter in ``data/facts.csv`` and grows
   forever. Quarter order comes from :func:`model.build.ordered_quarters`.
2. ``Snapshot`` gains a QoQ anchor. Three change rows are rendered, all on the
   run-rate basis: year-over-year (rolls forward), sequential (rolls forward) and
   versus the permanent baseline quarter (fixed). Anchors that do not exist yet
   render as ``n/a``, never as a fabricated number.

Usage
-----
    python scripts/build_workbook.py                 # data/ -> build/
    python scripts/build_workbook.py --data-dir <d> --out-dir <o>

The frozen v02 workbook is never opened for writing and the script refuses to
write to its path.
"""

from __future__ import annotations

import argparse
import datetime as _dt
import hashlib
import re
import sys
import zipfile
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import openpyxl
import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

if hasattr(sys.stdout, "reconfigure"):  # the workbook carries x, delta, en-dashes
    sys.stdout.reconfigure(encoding="utf-8")

REPO = Path(__file__).resolve().parents[1]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from model import build as mbuild  # noqa: E402
from model.calc import ProxyBasis, Scenario  # noqa: E402

__all__ = [
    "GENERATOR_VERSION",
    "FROZEN_WORKBOOK",
    "DataLayer",
    "build_workbook",
    "output_path_for",
    "main",
]

# --------------------------------------------------------------------------- #
# Constants
# --------------------------------------------------------------------------- #

#: Lineage tag stamped into the output filename. v02 is the frozen workbook;
#: everything this renderer produces is v03 and up.
GENERATOR_VERSION = "v03"

FROZEN_WORKBOOK = REPO / "ai_capex_forward_roic_analysis_v02.xlsx"
DEFAULT_DATA_DIR = REPO / "data"
DEFAULT_OUT_DIR = REPO / "build"

#: Company order on the presentation sheets.
TICKERS = mbuild.TICKERS

#: The quarter series frozen into v02. Used ONLY to locate a note/hyperlink's
#: original address so it can be carried forward; it never bounds the output.
V02_QUARTERS = mbuild.QUARTERS

SHEET_ORDER = ("Trajectory", "Snapshot", "Inputs", "Sources & Notes", "Checks")

# --- palette ---------------------------------------------------------------
RGB_HEADER = "FF1F4E78"
RGB_FACT = "FFEAF2F8"  # blue: sourced filing fact
RGB_ASSUMPTION = "FFFFF2CC"  # yellow: analyst assumption
RGB_DIVIDER = "FFD9E2F3"
RGB_GREY = "FFF2F2F2"
RGB_WHITE = "FFFFFFFF"
RGB_GRID = "FFD9D9D9"
RGB_LINK = "FF0000FF"

FILL_HEADER = PatternFill(fill_type="solid", start_color=RGB_HEADER, end_color=RGB_HEADER)
FILL_FACT = PatternFill(fill_type="solid", start_color=RGB_FACT, end_color=RGB_FACT)
FILL_ASSUMPTION = PatternFill(
    fill_type="solid", start_color=RGB_ASSUMPTION, end_color=RGB_ASSUMPTION
)
FILL_DIVIDER = PatternFill(fill_type="solid", start_color=RGB_DIVIDER, end_color=RGB_DIVIDER)
FILL_GREY = PatternFill(fill_type="solid", start_color=RGB_GREY, end_color=RGB_GREY)
FILL_WHITE = PatternFill(fill_type="solid", start_color=RGB_WHITE, end_color=RGB_WHITE)

FONT_TITLE = Font(name="Arial", size=10, bold=True, color=RGB_WHITE)
FONT_BANNER = Font(name="Arial", size=9, italic=True)
FONT_BODY = Font(name="Arial", size=9, color="FF000000")
FONT_BODY_BOLD = Font(name="Arial", size=9, bold=True)
FONT_LINK = Font(name="Arial", size=9, color=RGB_LINK)
FONT_LINK_UNDERLINE = Font(name="Arial", size=9, color=RGB_LINK, underline="single")

_SIDE = Side(style="thin", color=RGB_GRID)
BORDER_CELL = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)
BORDER_HEADER = Border(bottom=Side(style="medium", color=RGB_HEADER))

ALIGN_HEADER = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_BODY = Alignment(vertical="top", wrap_text=True)
ALIGN_BANNER = Alignment(wrap_text=True)

# --- number formats (verbatim from v02) ------------------------------------
NF_GENERAL = "General"
NF_DATE = r"mmm\ d\,\ yyyy"
NF_USD1 = r"\$0.0"
NF_USD3 = r"\$0.000"
NF_USD4 = r"\$0.0000"
NF_PCT1 = "0.0%"
NF_PCT4 = "0.0000%"
NF_DEC1 = "0.0"
NF_DEC2 = "0.00"
NF_BPS = r"\+0;\-0;0"

# --- prose carried over from v02 -------------------------------------------
TRAJECTORY_TITLE = (
    "AI Capex / Revenue Proxy and Forward ROIC — Filing-Based Reconstruction"
)
TRAJECTORY_BANNER = (
    "Public facts are linked in notes. AI attribution, duration and margins are "
    "assumptions. Spread is shown in percentage points; only the change rows on "
    "Snapshot are basis points."
)
SNAPSHOT_TITLE = "Latest-Quarter Forward ROIC Snapshot — Filing-Based Reconstruction"
INPUTS_TITLE_LEFT = "Filing Inputs"
INPUTS_TITLE_RIGHT = "Model Assumptions and Latest Annual Capex Denominators"
INPUTS_BANNER = (
    "Blue cells are public-filing or official-company facts. Yellow cells are "
    "analyst assumptions. Every sourced value carries a cell note with the public URL."
)
SOURCES_TITLE = "Public Source Ledger and Model Caveats"
SOURCES_BANNER = (
    "Source URLs are public. BamSEC was not available in this task, so SEC EDGAR and "
    "official company investor-relations disclosures are the authority. Model "
    "assumptions are explicitly separated from facts."
)
CHECKS_TITLE = "Independent Recalculation Checks"
CAVEATS_TITLE = "Critical Interpretation and Comparability Caveats"

#: The seven caveat bullets. Bullet 2 states a fact about the sheet itself and is
#: templated on the quarter count; the other six are v02 prose, verbatim.
CAVEAT_BULLETS: tuple[str, ...] = (
    "• This is a proxy model, not an accounting ROIC calculation. RPO/backlog is "
    "converted into an annualized revenue proxy using assumed AI attribution and duration.",
    "• {n_quarters} quarters are shown and nothing rolls off. The change rows report "
    "the year-over-year, sequential (QoQ) and baseline comparisons of the run-rate spread. "
    "The original model assumptions are held constant to isolate reported-financial movement.",
    "• RPO/backlog is not like-for-like: Microsoft uses commercial RPO; Alphabet's "
    "definition changed in Q1 2026 and now includes TPU system-sale agreements; Amazon "
    "reports primarily AWS commitments; Oracle reports company-wide RPO.",
    "• Current disclosed durations differ from the retained assumptions: Microsoft "
    "disclosed 2.3 years versus the model's 2.5; Amazon disclosed 6.4 years versus the "
    "model's 4.0. These differences can materially change the revenue proxy.",
    "• Capex definitions differ by issuer. Microsoft uses a management measure affected "
    "by lease classification; Alphabet and Oracle use cash PP&E/capex; Amazon uses gross "
    "productive-asset cash payments; Meta includes finance-lease principal.",
    "• Microsoft's $175B reported capex outlook is lower because leases shifted outside "
    "its metric, not because economic investment plans fell. Alphabet and Meta denominators "
    "are range midpoints. Oracle uses FY2026 gross-capex actual because its approximately "
    "$70B FY2027 outlook is a non-comparable net-cash-outlay measure.",
    "• AI shares of RPO/revenue and capex, NOPAT margins, duration assumptions, and "
    "Damodaran sector mappings are analyst assumptions. Spread is percentage points; only "
    "the change rows are basis points.",
)

#: Trajectory row-1 label per company. Presentation chrome, not model logic; the
#: workbook names each issuer's headline quantity differently.
TRAJECTORY_FACT_LABEL: dict[str, str] = {
    "MSFT": "Commercial RPO/Backlog ($B)",
    "GOOG": "Revenue Backlog/RPO ($B)",
    "AMZN": "AWS-linked Commitments/RPO ($B)",
    "ORCL": "RPO ($B)",
    "META": "Revenue (Quarter, $B)",
}

TRAJECTORY_ROW_LABELS: tuple[str, ...] = (
    "",  # replaced per company by TRAJECTORY_FACT_LABEL
    "Quarterly Capex ($B)",
    "Annualized AI Capex ($B)",
    "Annualized AI Revenue Proxy ($B)",
    "Forward ROIC (Base)",
    "Spread vs WACC (ppt)",
)

#: Trajectory block: metric-row offset -> (DataFrame column, number format).
TRAJECTORY_ROWS: tuple[tuple[str, str], ...] = (
    ("fact_value_b", NF_USD1),
    ("quarterly_capex_b", NF_USD1),
    ("ai_capex_b", NF_USD1),
    ("ai_revenue_proxy_b", NF_USD1),
    ("forward_roic", NF_PCT1),
    ("spread", NF_PCT1),
)

TRAJECTORY_BLOCK_HEIGHT = 7  # 6 metric rows + 1 spacer

# Snapshot row numbers. Rows 4-24 are v02's own layout, unchanged. Rows 25-31 are
# the trajectory-view block, widened from three rows to seven by the QoQ anchor.
SNAP_ROW = {
    "latest_quarter": 4,
    "revenue_source": 5,
    "fact": 6,
    "ai_share": 7,
    "ai_linked": 8,
    "capex_guide": 9,
    "plan_basis": 10,
    "ai_share_capex": 11,
    "ai_capex": 12,
    "duration": 13,
    "revenue_proxy": 14,
    "nopat_base": 15,
    "divider_snapshot": 16,
    "roic_base": 17,
    "wacc": 18,
    "spread_base": 19,
    "roic_bear": 20,
    "spread_bear": 21,
    "roic_bull": 22,
    "spread_bull": 23,
    "divider_trajectory": 24,
    "rr_latest": 25,
    "rr_yoy": 26,
    "rr_qoq": 27,
    "rr_baseline": 28,
    "delta_yoy": 29,
    "delta_qoq": 30,
    "delta_baseline": 31,
}

#: v02 Snapshot row -> semantic key, for locating the note to carry forward.
#: v02 rows 25/26/27 map to rr_latest / rr_baseline / delta_baseline: the workbook's
#: single change row was the baseline comparison.
V02_SNAP_ROW_KEY = {
    4: "latest_quarter",
    5: "revenue_source",
    6: "fact",
    7: "ai_share",
    8: "ai_linked",
    9: "capex_guide",
    10: "plan_basis",
    11: "ai_share_capex",
    12: "ai_capex",
    13: "duration",
    14: "revenue_proxy",
    15: "nopat_base",
    17: "roic_base",
    18: "wacc",
    19: "spread_base",
    20: "roic_bear",
    21: "spread_bear",
    22: "roic_bull",
    23: "spread_bull",
    25: "rr_latest",
    26: "rr_baseline",
    27: "delta_baseline",
}
V02_SNAP_KEY_ROW = {v: k for k, v in V02_SNAP_ROW_KEY.items()}

#: Inputs fact-table columns, in sheet order.
INPUTS_FACT_COLUMNS: tuple[tuple[str, str, str], ...] = (
    # (column letter, facts.csv field, number format)
    ("A", "company", NF_GENERAL),
    ("B", "ticker", NF_GENERAL),
    ("C", "report_bucket", NF_GENERAL),
    ("D", "fiscal_period", NF_GENERAL),
    ("E", "period_end", NF_DATE),
    ("F", "rpo_backlog_or_revenue_usd_b", NF_USD3),
    ("G", "fact_metric", NF_GENERAL),
    ("H", "quarterly_capex_usd_b", NF_USD3),
    ("I", "capex_definition", NF_GENERAL),
    ("J", "fact_source_url", NF_GENERAL),
    ("K", "capex_source_url", NF_GENERAL),
    ("L", "evidence_derivation", NF_GENERAL),
    ("M", "fact_source_id", NF_GENERAL),
    ("N", "capex_source_id", NF_GENERAL),
)

INPUTS_FACT_HEADERS = (
    "Company",
    "Ticker",
    "Report Bucket",
    "Fiscal Period",
    "Period End",
    "RPO/Backlog or Revenue ($B)",
    "Fact Metric",
    "Quarterly Capex ($B)",
    "Capex Definition",
    "Fact Source URL",
    "Capex Source URL",
    "Evidence / Derivation",
    "Fact Source ID",
    "Capex Source ID",
)

#: Inputs assumption-block columns: (letter, assumptions.csv field, nf, fill class).
INPUTS_ASSUMPTION_COLUMNS: tuple[tuple[str, str, str, str], ...] = (
    ("P", "ticker", NF_GENERAL, "other"),
    ("Q", "ai_revenue_proxy", NF_GENERAL, "other"),
    ("R", "ai_share_of_rpo_revenue", NF_PCT1, "assumption"),
    ("S", "rpo_duration_years", NF_DEC1, "assumption"),
    ("T", "ai_share_of_capex", NF_PCT1, "assumption"),
    ("U", "nopat_margin_bear", NF_PCT1, "assumption"),
    ("V", "nopat_margin_base", NF_PCT1, "assumption"),
    ("W", "nopat_margin_bull", NF_PCT1, "assumption"),
    ("X", "wacc", NF_PCT1, "fact"),
    ("Y", "damodaran_sector_date", NF_GENERAL, "other"),
    ("Z", "annual_capex_guide_midpoint_actual_usd_b", NF_USD1, "fact"),
    ("AA", "plan_basis", NF_GENERAL, "other"),
    ("AB", "plan_source_url", NF_GENERAL, "fact"),
    ("AC", "source_assumption_caveat", NF_GENERAL, "other"),
)

INPUTS_ASSUMPTION_HEADERS = (
    "Ticker",
    "AI Revenue Proxy",
    "AI Share of RPO/Revenue",
    "RPO Duration (Years)",
    "AI Share of Capex",
    "NOPAT Margin (Bear)",
    "NOPAT Margin (Base)",
    "NOPAT Margin (Bull)",
    "WACC",
    "Damodaran Sector / Date",
    "Annual Capex Guide / Midpoint / Actual ($B)",
    "Plan Basis",
    "Plan Source URL",
    "Source / Assumption Caveat",
)

SOURCES_HEADERS = (
    "Source ID",
    "Company",
    "Period",
    "Metric",
    "Value",
    "Classification",
    "Filing / Disclosure",
    "Public URL",
    "Evidence / Derivation",
    "Local Source",
    "Status",
    "Caveat",
)

CHECKS_HEADERS = ("Check ID", "Test", "Expected", "Workbook Value", "Tolerance", "Status")

#: Column widths, verbatim from v02. Quarter columns beyond the first keep Excel's
#: default width, exactly as v02 left D:G.
COLUMN_WIDTHS = {
    "Trajectory": {"A": 13.0, "B": 39.0, "C": 17.0},
    "Snapshot": {"A": 53.0, "B": 25.0},
    "Inputs": {
        "A": 17.0, "B": 9.0, "C": 12.0, "D": 14.0, "E": 13.0, "F": 19.0, "G": 29.0,
        "H": 18.0, "I": 42.0, "L": 65.0, "M": 20.0, "N": 21.0, "P": 9.0, "Q": 39.0,
        "R": 18.0, "S": 17.0, "T": 16.0, "U": 17.0, "X": 12.0, "Y": 44.0, "Z": 21.0,
        "AA": 37.0, "AB": 44.0, "AC": 70.0,
    },
    "Sources & Notes": {
        "A": 20.0, "B": 18.0, "C": 22.0, "D": 33.0, "E": 14.0, "F": 35.0, "G": 30.0,
        "H": 48.0, "I": 85.0, "J": 55.0, "K": 24.0, "L": 60.0,
    },
    "Checks": {"A": 42.0, "B": 48.0, "C": 20.0, "E": 14.0, "F": 12.0},
}

COMMENT_WIDTH, COMMENT_HEIGHT = 144, 79
COMMENT_AUTHOR = "Codex"

#: Fixed document properties, so the output is reproducible.
DOC_PROPERTIES = {
    "creator": "Brookfield Global Strategy Group / Codex",
    "title": "AI Capex / Revenue Proxy and Forward ROIC Analysis",
    "subject": "Filing-based reconstruction of the user-provided screenshots",
    "description": (
        "Values-only render of the data layer: public filing inputs with linked cell "
        "notes, explicit model assumptions, computed trajectory and snapshot views, "
        "and independent checks."
    ),
    "lastModifiedBy": "scripts/build_workbook.py",
}
FIXED_TIMESTAMP = _dt.datetime(2026, 1, 1, 0, 0, 0)
ZIP_TIMESTAMP = (1980, 1, 1, 0, 0, 0)


# --------------------------------------------------------------------------- #
# Data layer
# --------------------------------------------------------------------------- #


def _clean(value: Any) -> Any:
    """NaN/NaT -> None; everything else unchanged."""
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def _text(value: Any) -> str:
    v = _clean(value)
    return "" if v is None else str(v)


class DataLayer:
    """Everything read off ``data/``, indexed for rendering."""

    def __init__(self, data_dir: Path | str = DEFAULT_DATA_DIR) -> None:
        self.data_dir = Path(data_dir)
        self.facts_df = pd.read_csv(self.data_dir / "facts.csv")
        self.assumptions_df = pd.read_csv(self.data_dir / "assumptions.csv")
        self.sources_df = pd.read_csv(self.data_dir / "sources.csv")
        self.notes_df = pd.read_csv(self.data_dir / "cell_notes.csv")
        self.links_df = pd.read_csv(self.data_dir / "hyperlinks.csv")

        self.facts, self.assumptions = mbuild.load_inputs_from_csv(self.data_dir)
        self.quarters = mbuild.ordered_quarters(self.facts)

        unknown = sorted({f.ticker for f in self.facts} - set(TICKERS))
        if unknown:
            raise ValueError(
                "facts.csv contains tickers the presentation layout does not know: "
                + ", ".join(unknown)
            )

        # facts.csv rows keyed by (ticker, quarter)
        self.fact_rows: dict[tuple[str, str], Mapping[str, Any]] = {
            (str(r["ticker"]), str(r["report_bucket"])): r
            for _, r in self.facts_df.iterrows()
        }
        # the applicable assumption row per ticker (same selection model/build uses)
        newest = max(f.period_end for f in self.facts)
        assum = self.assumptions_df.assign(
            _eff=pd.to_datetime(self.assumptions_df["effective_from"])
        ).sort_values("_eff")
        self.assumption_rows: dict[str, Mapping[str, Any]] = {}
        for ticker, group in assum.groupby("ticker", sort=False):
            applicable = group[group["_eff"] <= pd.Timestamp(newest)]
            self.assumption_rows[str(ticker)] = (
                applicable if not applicable.empty else group
            ).iloc[-1]

        self.sources: dict[str, Mapping[str, Any]] = {
            str(r["source_id"]): r for _, r in self.sources_df.iterrows()
        }
        self.ledger: list[Mapping[str, Any]] = [
            r for _, r in self.sources_df.iterrows()
            if str(r.get("in_workbook_ledger", "yes")).strip().lower() == "yes"
        ]

        # notes and hyperlinks keyed by their v02 address
        self.notes: dict[tuple[str, str], str] = {
            (str(r["sheet"]), str(r["cell"])): str(r["note_text"])
            for _, r in self.notes_df.iterrows()
        }
        self.links: dict[tuple[str, str], Mapping[str, Any]] = {
            (str(r["sheet"]), str(r["cell"])): r for _, r in self.links_df.iterrows()
        }

    # -- convenience ------------------------------------------------------- #

    def fact(self, ticker: str, quarter: str) -> Mapping[str, Any]:
        return self.fact_rows[(ticker, quarter)]

    def assumption(self, ticker: str) -> Mapping[str, Any]:
        return self.assumption_rows[ticker]

    def wacc_url(self, ticker: str) -> str:
        row = self.sources.get(f"{ticker}-WACC")
        return _text(row["url"]) if row is not None else ""

    def source_evidence(self, source_id: str) -> str:
        row = self.sources.get(str(source_id))
        return _text(row["evidence_derivation"]) if row is not None else ""

    def source_local_path(self, source_id: str) -> str:
        row = self.sources.get(str(source_id))
        return _text(row["local_path_if_any"]) if row is not None else ""


# --------------------------------------------------------------------------- #
# Note carry-forward
# --------------------------------------------------------------------------- #

_REWRITABLE_PREFIXES = (
    "Value: ",
    "Public source: ",
    "Evidence: ",
    "Local source: ",
    "Fact source: ",
    "Capex source: ",
    "Contextual public filing: ",
    "Formula: ",
)


def rewrite_note(template: str, **fields: str | None) -> str:
    """Rewrite a v02 note line-for-line with new values, keeping every other line.

    ``fields`` keys are ``header``, ``value``, ``public_source``, ``evidence``,
    ``local_source``, ``fact_source``, ``capex_source``, ``contextual``, ``formula``.
    A field left out means "keep the template's line". ``local_source=""`` removes
    the line; a ``local_source`` on a template that has none inserts it after the
    Evidence line, which is where v02 puts it.

    Every substituted value comes from the data layer; nothing is invented. For a
    quarter that already existed in v02 all substitutions are no-ops, which is
    asserted by the parity test.
    """
    replacements = {
        "Value: ": fields.get("value"),
        "Public source: ": fields.get("public_source"),
        "Evidence: ": fields.get("evidence"),
        "Local source: ": fields.get("local_source"),
        "Fact source: ": fields.get("fact_source"),
        "Capex source: ": fields.get("capex_source"),
        "Contextual public filing: ": fields.get("contextual"),
        "Formula: ": fields.get("formula"),
    }
    lines = template.split("\n")
    out: list[str] = []
    seen_local = False
    for index, line in enumerate(lines):
        if index == 0 and fields.get("header") is not None:
            out.append(str(fields["header"]))
            continue
        matched = False
        for prefix in _REWRITABLE_PREFIXES:
            if not line.startswith(prefix):
                continue
            matched = True
            replacement = replacements.get(prefix)
            if prefix == "Local source: ":
                seen_local = True
                if replacement is None:
                    out.append(line)
                elif replacement:
                    out.append(prefix + replacement)
                # empty string -> drop the line
            elif replacement is None:
                out.append(line)
            else:
                out.append(prefix + replacement)
            break
        if matched:
            continue
        out.append(line)
        if line.startswith("Evidence: ") and fields.get("local_source") and not seen_local:
            out.append("Local source: " + str(fields["local_source"]))
            seen_local = True
    return "\n".join(out)


class NoteIndex:
    """Locate a v02 note by what it *means*, not by where it happened to sit.

    v02's ``Inputs`` sheet lays companies out in blocks of five quarter rows, so
    appending a sixth quarter shifts every row below Microsoft's block. Keying by
    ``(ticker, quarter, column)`` instead of by address means the audit trail
    survives that shift.
    """

    def __init__(self, data: DataLayer) -> None:
        self._data = data
        self._v02_quarters = tuple(q for q in V02_QUARTERS if q in data.quarters)
        if not self._v02_quarters:
            raise ValueError("no v02 quarter survives in facts.csv; cannot carry notes")
        self._n_v02 = len(V02_QUARTERS)
        self._last_v02 = self._v02_quarters[-1]

        # source_id -> v02 Sources & Notes row (ledger order is the sheet order)
        self._source_row: dict[str, int] = {}
        for index, row in enumerate(data.ledger):
            self._source_row[str(row["source_id"])] = 5 + index

    # -- v02 address lookups ---------------------------------------------- #

    def _q_index(self, quarter: str) -> int | None:
        try:
            return V02_QUARTERS.index(quarter)
        except ValueError:
            return None

    def trajectory(self, ticker: str, quarter: str, offset: int) -> str | None:
        """Note template for Trajectory metric row ``offset`` of a company-quarter."""
        q = self._q_index(quarter)
        if q is None:
            q = self._q_index(self._last_v02)
        col = "CDEFG"[q]
        row = 5 + TRAJECTORY_BLOCK_HEIGHT * TICKERS.index(ticker) + offset
        return self._data.notes.get(("Trajectory", f"{col}{row}"))

    def inputs_fact(self, ticker: str, quarter: str, column: str) -> str | None:
        q = self._q_index(quarter)
        if q is None:
            q = self._q_index(self._last_v02)
        row = 5 + self._n_v02 * TICKERS.index(ticker) + q
        return self._data.notes.get(("Inputs", f"{column}{row}"))

    def inputs_assumption(self, ticker: str, column: str) -> str | None:
        row = 5 + TICKERS.index(ticker)
        return self._data.notes.get(("Inputs", f"{column}{row}"))

    def snapshot(self, ticker: str, key: str) -> str | None:
        row = V02_SNAP_KEY_ROW.get(key)
        if row is None:
            return None
        col = "BCDEF"[TICKERS.index(ticker)]
        return self._data.notes.get(("Snapshot", f"{col}{row}"))

    def source(self, source_id: str) -> str | None:
        row = self._source_row.get(str(source_id))
        if row is None:
            return None
        return self._data.notes.get(("Sources & Notes", f"H{row}"))


# --------------------------------------------------------------------------- #
# Cell helpers
# --------------------------------------------------------------------------- #


def _put(
    ws,
    coordinate: str,
    value: Any,
    *,
    font: Font = FONT_BODY,
    fill: PatternFill | None = None,
    number_format: str = NF_GENERAL,
    alignment: Alignment = ALIGN_BODY,
    border: Border | None = BORDER_CELL,
    note: str | None = None,
    link: str | None = None,
    link_display: bool = False,
):
    cell = ws[coordinate]
    cell.value = value
    cell.font = font
    if fill is not None:
        cell.fill = fill
    cell.number_format = number_format
    cell.alignment = alignment
    if border is not None:
        cell.border = border
    if note:
        comment = Comment(note, COMMENT_AUTHOR, height=COMMENT_HEIGHT, width=COMMENT_WIDTH)
        cell.comment = comment
    if link:
        cell.hyperlink = Hyperlink(
            ref=coordinate, target=link, display=link if link_display else None
        )
    return cell


def _title(ws, coordinate: str, text: str) -> None:
    _put(
        ws,
        coordinate,
        text,
        font=FONT_TITLE,
        fill=FILL_HEADER,
        alignment=ALIGN_HEADER,
        border=BORDER_HEADER,
    )


def _banner(ws, coordinate: str, text: str, fill: PatternFill) -> None:
    _put(
        ws,
        coordinate,
        text,
        font=FONT_BANNER,
        fill=fill,
        alignment=ALIGN_BANNER,
        border=None,
    )


def _header(ws, row: int, first_col: int, labels: Iterable[str]) -> None:
    for offset, label in enumerate(labels):
        _put(
            ws,
            f"{get_column_letter(first_col + offset)}{row}",
            label,
            font=FONT_TITLE,
            fill=FILL_HEADER,
            alignment=ALIGN_HEADER,
            border=BORDER_HEADER,
        )


def _apply_sheet_chrome(ws, name: str) -> None:
    ws.sheet_view.showGridLines = False
    for column, width in COLUMN_WIDTHS.get(name, {}).items():
        ws.column_dimensions[column].width = width


# --------------------------------------------------------------------------- #
# Trajectory
# --------------------------------------------------------------------------- #


def _build_trajectory(ws, data: DataLayer, notes: NoteIndex, trajectory: pd.DataFrame) -> None:
    quarters = data.quarters
    last_col = 2 + len(quarters)
    last_letter = get_column_letter(last_col)

    _title(ws, "A1", TRAJECTORY_TITLE)
    ws.merge_cells(f"A1:{last_letter}1")
    _banner(ws, "A2", TRAJECTORY_BANNER, FILL_ASSUMPTION)
    ws.merge_cells(f"A2:{last_letter}2")
    ws.row_dimensions[1].height = 25.0
    ws.row_dimensions[2].height = 36.0

    _header(ws, 4, 1, ("Company", "Metric", *quarters))

    indexed = trajectory.set_index(["ticker", "quarter"])
    for company_index, ticker in enumerate(TICKERS):
        top = 5 + TRAJECTORY_BLOCK_HEIGHT * company_index
        for offset, (field, number_format) in enumerate(TRAJECTORY_ROWS):
            row = top + offset
            if offset == 0:
                _put(ws, f"A{row}", ticker)
            else:
                _put(ws, f"A{row}", None)
            label = (
                TRAJECTORY_FACT_LABEL[ticker]
                if offset == 0
                else TRAJECTORY_ROW_LABELS[offset]
            )
            _put(
                ws,
                f"B{row}",
                label,
                font=FONT_BODY_BOLD if offset >= 4 else FONT_BODY,
            )
            for quarter_index, quarter in enumerate(quarters):
                coordinate = f"{get_column_letter(3 + quarter_index)}{row}"
                value = float(indexed.loc[(ticker, quarter), field])
                fact_row = data.fact(ticker, quarter)
                sourced = offset in (0, 1)
                if offset == 0:
                    link = _text(fact_row["fact_source_url"])
                    note_ctx = _fact_note_context(data, ticker, quarter, kind="fact")
                elif offset == 1:
                    link = _text(fact_row["capex_source_url"])
                    note_ctx = _fact_note_context(data, ticker, quarter, kind="capex")
                else:
                    link = None
                    note_ctx = _derived_note_context(data, ticker, quarter)
                template = notes.trajectory(ticker, quarter, offset)
                note = rewrite_note(template, **note_ctx) if template else None
                _put(
                    ws,
                    coordinate,
                    value,
                    font=FONT_LINK if sourced else FONT_BODY,
                    fill=FILL_FACT if sourced else None,
                    number_format=number_format,
                    note=note,
                    link=link,
                    link_display=True,
                )
        # spacer row keeps the v02 block rhythm
    ws.freeze_panes = "C4"
    ws.page_setup.orientation = "landscape"


def _fact_note_context(
    data: DataLayer, ticker: str, quarter: str, *, kind: str
) -> dict[str, str]:
    """Line replacements for a blue fact/capex cell's note."""
    row = data.fact(ticker, quarter)
    if kind == "fact":
        header = f"{ticker} {quarter} — {_text(row['fact_metric'])}"
        value = f"${float(row['rpo_backlog_or_revenue_usd_b']):.3f}B"
        url = _text(row["fact_source_url"])
        source_id = _text(row["fact_source_id"])
    else:
        header = f"{ticker} {quarter} — quarterly capex"
        value = f"${float(row['quarterly_capex_usd_b']):.3f}B"
        url = _text(row["capex_source_url"])
        source_id = _text(row["capex_source_id"])
    return {
        "header": header,
        "value": value,
        "public_source": url,
        "evidence": data.source_evidence(source_id),
        "local_source": data.source_local_path(source_id),
    }


def _derived_note_context(data: DataLayer, ticker: str, quarter: str) -> dict[str, str]:
    row = data.fact(ticker, quarter)
    return {
        "header": f"DERIVED MODEL OUTPUT — {ticker} {quarter}",
        "fact_source": _text(row["fact_source_url"]),
        "capex_source": _text(row["capex_source_url"]),
    }


# --------------------------------------------------------------------------- #
# Snapshot
# --------------------------------------------------------------------------- #


def _snapshot_labels(
    yoy: str | None, qoq: str | None, baseline: str, latest_quarter: str
) -> dict[str, str]:
    """Row labels for the trajectory-view block, naming the comparison quarter.

    ``yoy`` and ``qoq`` are ``None`` when the series is too short to reach that
    anchor; the label says so rather than implying a number exists.
    """
    return {
        "rr_latest": f"Spread {latest_quarter} (run-rate basis, ppt)",
        "rr_yoy": (
            f"Spread {yoy} (run-rate basis, ppt) — YoY anchor"
            if yoy
            else "Spread (YoY anchor) — not yet available"
        ),
        "rr_qoq": (
            f"Spread {qoq} (run-rate basis, ppt) — QoQ anchor"
            if qoq
            else "Spread (QoQ anchor) — not yet available"
        ),
        "rr_baseline": f"Spread {baseline} (run-rate basis, ppt) — baseline anchor",
        "delta_yoy": (
            f"Δ Spread {latest_quarter} vs {yoy} (YoY, bps)"
            if yoy
            else "Δ Spread (YoY, bps) — not yet available"
        ),
        "delta_qoq": (
            f"Δ Spread {latest_quarter} vs {qoq} (QoQ, bps)"
            if qoq
            else "Δ Spread (QoQ, bps) — not yet available"
        ),
        "delta_baseline": f"Δ Spread {latest_quarter} vs {baseline} (baseline, bps)",
    }


def _build_snapshot(
    ws,
    data: DataLayer,
    notes: NoteIndex,
    snapshot: pd.DataFrame,
    latest_quarter: str,
    baseline_quarter: str,
) -> None:
    _title(ws, "A1", SNAPSHOT_TITLE)
    ws.merge_cells("A1:F1")
    _header(ws, 3, 1, ("Metric", *TICKERS))
    for row in (4, 5, 10):
        ws.row_dimensions[row].height = 34.0

    base_rows = snapshot[snapshot["scenario"] == Scenario.BASE.value].set_index("ticker")
    bear_rows = snapshot[snapshot["scenario"] == Scenario.BEAR.value].set_index("ticker")
    bull_rows = snapshot[snapshot["scenario"] == Scenario.BULL.value].set_index("ticker")

    first = base_rows.loc[TICKERS[0]]
    labels = _snapshot_labels(
        _clean(first["yoy_quarter"]),
        _clean(first["qoq_quarter"]),
        baseline_quarter,
        latest_quarter,
    )

    # column A labels
    static_labels = {
        "latest_quarter": "Latest Quarter",
        "revenue_source": "AI Revenue Source",
        "fact": "Total RPO/Backlog or Annualized Revenue ($B)",
        "ai_share": "AI Share % (assumption)",
        "ai_linked": "AI-linked RPO/Revenue ($B)",
        "capex_guide": "Annual Capex Guide / Midpoint / Actual ($B)",
        "plan_basis": "Plan Basis",
        "ai_share_capex": "AI Share of Capex % (assumption)",
        "ai_capex": "AI Capex ($B)",
        "duration": "RPO Duration (years, assumption)",
        "revenue_proxy": "Annualized AI Revenue Proxy ($B)",
        "nopat_base": "NOPAT Margin (Base)",
        "roic_base": "Forward ROIC (Base)",
        "wacc": "WACC (Damodaran sector, Jan. 2026)",
        "spread_base": "Spread (Base, ppt)",
        "roic_bear": "Forward ROIC (Bear)",
        "spread_bear": "Spread (Bear, ppt)",
        "roic_bull": "Forward ROIC (Bull)",
        "spread_bull": "Spread (Bull, ppt)",
    }
    static_labels.update(labels)

    for key, label in static_labels.items():
        _put(ws, f"A{SNAP_ROW[key]}", label)

    # section dividers
    for key, text in (
        ("divider_snapshot", "— SNAPSHOT VIEW (latest annual capex denominator) —"),
        (
            "divider_trajectory",
            "— TRAJECTORY VIEW (quarterly capex × 4 as denominator) —",
        ),
    ):
        row = SNAP_ROW[key]
        _put(ws, f"A{row}", text, font=FONT_BODY_BOLD, fill=FILL_DIVIDER)
        for column in "BCDEF":
            _put(ws, f"{column}{row}", None, fill=FILL_DIVIDER)

    for index, ticker in enumerate(TICKERS):
        column = "BCDEF"[index]
        base = base_rows.loc[ticker]
        bear = bear_rows.loc[ticker]
        bull = bull_rows.loc[ticker]
        assumption = data.assumption(ticker)
        fact_row = data.fact(ticker, latest_quarter)
        is_revenue = ProxyBasis(base["proxy_basis"]) is ProxyBasis.REVENUE

        # On the Snapshot the capex denominator is the ANNUAL plan, so the note's
        # "Capex source" line points at the plan URL, not the quarter's capex filing.
        latest_ctx = {
            "fact_source": _text(fact_row["fact_source_url"]),
            "capex_source": _text(assumption["plan_source_url"]),
            "contextual": _text(fact_row["fact_source_url"]),
        }

        # v02 quirk, preserved: the two rows that sit between the fact and the
        # annual-plan denominator cite the quarter's capex filing; every other row
        # cites the plan.
        quarter_capex_ctx = {"capex_source": _text(fact_row["capex_source_url"])}

        def note_for(key: str, **extra: str) -> str | None:
            template = notes.snapshot(ticker, key)
            if not template:
                return None
            context = dict(latest_ctx)
            if key in ("ai_linked", "revenue_proxy"):
                context.update(quarter_capex_ctx)
            return rewrite_note(template, **{**context, **extra})

        def cell(key: str, value: Any, **kwargs: Any):
            return _put(ws, f"{column}{SNAP_ROW[key]}", value, **kwargs)

        period_end = base["period_end"]
        cell(
            "latest_quarter",
            f"{base['fiscal_period']} ({period_end:%b}. {period_end.day}, {period_end.year})",
        )
        cell("revenue_source", _text(assumption["ai_revenue_proxy"]))

        fact_value = float(fact_row["rpo_backlog_or_revenue_usd_b"])
        fact_value_line = f"${fact_value:.3f}B" + (" quarterly" if is_revenue else "")
        cell(
            "fact",
            float(base["snapshot_fact_b"]),
            font=FONT_LINK,
            fill=FILL_FACT,
            number_format=NF_USD1,
            link=_text(fact_row["fact_source_url"]),
            link_display=True,
            note=note_for(
                "fact",
                value=fact_value_line,
                public_source=_text(fact_row["fact_source_url"]),
                evidence=data.source_evidence(_text(fact_row["fact_source_id"])),
                local_source=data.source_local_path(_text(fact_row["fact_source_id"])),
            ),
        )
        cell(
            "ai_share",
            float(base["ai_share_of_fact"]),
            fill=FILL_ASSUMPTION,
            number_format=NF_PCT1,
            note=note_for("ai_share"),
        )
        cell(
            "ai_linked",
            float(base["ai_linked_b"]),
            number_format=NF_USD1,
            note=note_for("ai_linked"),
        )
        cell(
            "capex_guide",
            float(base["annual_capex_guide_b"]),
            font=FONT_LINK,
            fill=FILL_FACT,
            number_format=NF_USD1,
            link=_text(assumption["plan_source_url"]),
            link_display=True,
            note=note_for("capex_guide"),
        )
        cell("plan_basis", _text(base["plan_basis"]), note=note_for("plan_basis"))
        cell(
            "ai_share_capex",
            float(base["ai_share_of_capex"]),
            fill=FILL_ASSUMPTION,
            number_format=NF_PCT1,
            note=note_for("ai_share_capex"),
        )
        cell(
            "ai_capex",
            float(base["ai_capex_b"]),
            number_format=NF_USD1,
            note=note_for("ai_capex"),
        )
        duration = _clean(base["rpo_duration_years"])
        cell(
            "duration",
            "N/A" if duration is None else float(duration),
            fill=FILL_ASSUMPTION,
            number_format=NF_DEC1,
            note=note_for("duration"),
        )
        cell(
            "revenue_proxy",
            float(base["ai_revenue_proxy_b"]),
            number_format=NF_USD1,
            note=note_for("revenue_proxy"),
        )
        cell(
            "nopat_base",
            float(base["nopat_margin"]),
            fill=FILL_ASSUMPTION,
            number_format=NF_PCT1,
            note=note_for("nopat_base"),
        )
        cell(
            "roic_base",
            float(base["forward_roic"]),
            number_format=NF_PCT1,
            note=note_for("roic_base"),
        )
        cell(
            "wacc",
            float(base["wacc"]),
            font=FONT_LINK,
            fill=FILL_FACT,
            number_format=NF_PCT1,
            link=data.wacc_url(ticker),
            link_display=True,
            note=note_for("wacc"),
        )
        cell(
            "spread_base",
            float(base["spread"]),
            number_format=NF_PCT1,
            note=note_for("spread_base"),
        )
        for key, row_values in (
            ("roic_bear", bear["forward_roic"]),
            ("spread_bear", bear["spread"]),
            ("roic_bull", bull["forward_roic"]),
            ("spread_bull", bull["spread"]),
        ):
            cell(key, float(row_values), number_format=NF_PCT1, note=note_for(key))

        # -- trajectory-view block: run-rate levels then the three changes ---
        run_rate_cells = (
            ("rr_latest", base["spread_latest_runrate"], "rr_latest"),
            ("rr_yoy", base["spread_yoy_runrate"], "rr_latest"),
            ("rr_qoq", base["spread_qoq_runrate"], "rr_latest"),
            ("rr_baseline", base["spread_base_runrate"], "rr_baseline"),
        )
        for key, value, template_key in run_rate_cells:
            value = _clean(value)
            cell(
                key,
                "n/a" if value is None else float(value),
                number_format=NF_PCT1,
                note=note_for(template_key, formula=static_labels[key]),
            )
        delta_cells = (
            ("delta_yoy", base["delta_spread_yoy_bps"], "delta_baseline"),
            ("delta_qoq", base["delta_spread_qoq_bps"], "delta_baseline"),
            ("delta_baseline", base["delta_spread_bps"], "delta_baseline"),
        )
        for key, value, template_key in delta_cells:
            value = _clean(value)
            cell(
                key,
                "n/a" if value is None else float(value),
                number_format=NF_BPS,
                note=note_for(template_key, formula=static_labels[key]),
            )

    ws.freeze_panes = "B4"
    ws.page_setup.orientation = "landscape"


# --------------------------------------------------------------------------- #
# Inputs
# --------------------------------------------------------------------------- #


def _build_inputs(ws, data: DataLayer, notes: NoteIndex) -> None:
    quarters = data.quarters
    _title(ws, "A1", INPUTS_TITLE_LEFT)
    ws.merge_cells("A1:N1")
    _title(ws, "P1", INPUTS_TITLE_RIGHT)
    ws.merge_cells("P1:AC1")
    _banner(ws, "A2", INPUTS_BANNER, FILL_GREY)
    ws.merge_cells("A2:AC2")
    ws.row_dimensions[2].height = 34.0
    ws.row_dimensions[4].height = 26.0

    _header(ws, 4, 1, INPUTS_FACT_HEADERS)
    _header(ws, 4, 16, INPUTS_ASSUMPTION_HEADERS)  # column P

    link_columns = {
        "F": "fact_source_url",
        "H": "capex_source_url",
        "J": "fact_source_url",
        "K": "capex_source_url",
    }
    note_kind = {"F": "fact", "H": "capex", "J": "fact", "K": "capex"}

    row = 5
    for ticker in TICKERS:
        for quarter in quarters:
            fact_row = data.fact(ticker, quarter)
            ws.row_dimensions[row].height = 55.0
            for column, field, number_format in INPUTS_FACT_COLUMNS:
                value = _clean(fact_row[field])
                if field == "period_end" and value is not None:
                    value = _dt.datetime.fromisoformat(str(value))
                elif field in (
                    "rpo_backlog_or_revenue_usd_b",
                    "quarterly_capex_usd_b",
                ):
                    value = float(value)
                sourced = column in link_columns
                template = notes.inputs_fact(ticker, quarter, column)
                note = None
                if template:
                    note = rewrite_note(
                        template,
                        **_fact_note_context(
                            data, ticker, quarter, kind=note_kind[column]
                        ),
                    )
                _put(
                    ws,
                    f"{column}{row}",
                    value,
                    font=FONT_LINK_UNDERLINE if sourced else FONT_BODY,
                    fill=FILL_FACT if sourced else None,
                    number_format=number_format,
                    note=note,
                    link=_text(fact_row[link_columns[column]]) if sourced else None,
                    # Excel omits the display attribute where the cell text already
                    # IS the URL (columns J/K); it stores it where the cell shows a
                    # number (columns F/H). v02 does exactly this.
                    link_display=column in ("F", "H"),
                )
            row += 1

    fill_for = {
        "fact": FILL_FACT,
        "assumption": FILL_ASSUMPTION,
        "other": None,
    }
    for index, ticker in enumerate(TICKERS):
        assumption_row = 5 + index
        assumption = data.assumption(ticker)
        latest_fact = data.fact(ticker, quarters[-1])
        contextual = _text(latest_fact["fact_source_url"])
        for column, field, number_format, fill_class in INPUTS_ASSUMPTION_COLUMNS:
            value = _clean(assumption[field])
            if fill_class == "fact" and column in ("X", "Z"):
                value = None if value is None else float(value)
            elif fill_class == "assumption":
                value = None if value is None else float(value)
            template = notes.inputs_assumption(ticker, column)
            note = None
            if template:
                if template.startswith("MODEL ASSUMPTION"):
                    note = rewrite_note(template, contextual=contextual)
                else:
                    note = rewrite_note(template)
            link = None
            if column == "X":
                link = data.wacc_url(ticker)
            elif column in ("Z", "AB"):
                link = _text(assumption["plan_source_url"])
            _put(
                ws,
                f"{column}{assumption_row}",
                value,
                font=FONT_LINK_UNDERLINE if fill_class == "fact" else FONT_BODY,
                fill=fill_for[fill_class],
                number_format=number_format,
                note=note,
                link=link,
                link_display=column in ("X", "Z"),
            )

    ws.freeze_panes = "A5"


# --------------------------------------------------------------------------- #
# Sources & Notes
# --------------------------------------------------------------------------- #


def _split_title(title: str) -> tuple[str, str]:
    """``'msft-20250630.htm — Commercial RPO'`` -> ``('msft-20250630.htm', 'Commercial RPO')``."""
    marker = " — "
    if marker in title:
        filing, _, metric = title.partition(marker)
        return filing, metric
    return "", title


def _build_sources(ws, data: DataLayer, notes: NoteIndex, n_quarters: int) -> None:
    _title(ws, "A1", SOURCES_TITLE)
    ws.merge_cells("A1:L1")
    _banner(ws, "A2", SOURCES_BANNER, FILL_ASSUMPTION)
    ws.merge_cells("A2:L2")
    _header(ws, 4, 1, SOURCES_HEADERS)

    # The ledger's Value column is not one unit: dollar rows are $B, WACC rows are
    # decimals shown as a percentage. v02 formats each kind separately.
    value_format_by_kind = {
        "fact": NF_USD3,
        "capex": NF_USD3,
        "plan": NF_USD1,
        "wacc": "0.00%",
    }

    row = 5
    for entry in data.ledger:
        ws.row_dimensions[row].height = 52.0
        filing, metric = _split_title(_text(entry["title_or_description"]))
        reported = _clean(entry["reported_value"])
        value_format = value_format_by_kind.get(_text(entry["kind"]), NF_USD3)
        url = _text(entry["url"])
        source_id = _text(entry["source_id"])
        values = (
            ("A", source_id, NF_GENERAL),
            ("B", _text(entry["company"]), NF_GENERAL),
            ("C", _text(entry["period"]), NF_GENERAL),
            ("D", metric, NF_GENERAL),
            ("E", None if reported is None else float(reported), value_format),
            ("F", _text(entry["classification"]), NF_GENERAL),
            ("G", filing, NF_GENERAL),
            ("H", url, NF_GENERAL),
            ("I", _text(entry["evidence_derivation"]), NF_GENERAL),
            ("J", _text(entry["local_path_if_any"]), NF_GENERAL),
            ("K", _text(entry["status"]), NF_GENERAL),
            ("L", _text(entry["caveat"]), NF_GENERAL),
        )
        template = notes.source(source_id)
        note = None
        if template:
            note = rewrite_note(
                template,
                header=source_id,
                value=("" if reported is None else str(float(reported))),
                public_source=url,
                evidence=_text(entry["evidence_derivation"]),
                local_source=_text(entry["local_path_if_any"]),
            )
        for column, value, number_format in values:
            is_url = column == "H"
            is_value = column == "E"
            _put(
                ws,
                f"{column}{row}",
                value if value != "" else None,
                font=FONT_LINK_UNDERLINE if is_url else (FONT_LINK if is_value else FONT_BODY),
                fill=FILL_FACT if (is_url or is_value) else None,
                number_format=number_format,
                note=note if is_url else None,
                link=url if is_url else None,
            )
        row += 1

    caveat_row = row + 2
    _title(ws, f"A{caveat_row}", CAVEATS_TITLE)
    ws.merge_cells(f"A{caveat_row}:L{caveat_row}")
    for index, bullet in enumerate(CAVEAT_BULLETS):
        target = caveat_row + 1 + index
        _put(
            ws,
            f"A{target}",
            bullet.format(n_quarters=n_quarters),
            fill=FILL_WHITE if index % 2 == 0 else FILL_GREY,
        )
        for column in "BCDEFGHIJKL":
            _put(
                ws,
                f"{column}{target}",
                None,
                fill=FILL_WHITE if index % 2 == 0 else FILL_GREY,
            )
        ws.merge_cells(f"A{target}:L{target}")
        ws.row_dimensions[target].height = 38.0

    ws.freeze_panes = "A4"


# --------------------------------------------------------------------------- #
# Checks
# --------------------------------------------------------------------------- #


def _quarter_token(quarter: str) -> str:
    return quarter.replace(" ", "")


def _build_checks(
    wb,
    ws,
    data: DataLayer,
    trajectory: pd.DataFrame,
    snapshot: pd.DataFrame,
) -> list[tuple[str, str]]:
    """Compare the model's own output against what was actually rendered.

    In a values-only workbook there is nothing for Excel to recompute, so these are
    written as values with a literal PASS/FAIL rather than as live formulas: column
    ``Expected`` is the number ``model/calc.py`` produced, column ``Workbook Value``
    is read back out of the rendered cell. That catches a mis-placed or mis-typed
    render -- the failure mode a values-only sheet actually has -- and stays
    readable without a spreadsheet engine. It is deliberately NOT an arithmetic
    re-derivation: there is one calculation engine, by design.

    Returns the list of failures as ``(check_id, message)``.
    """
    _title(ws, "A1", CHECKS_TITLE)
    ws.merge_cells("A1:F1")
    _header(ws, 3, 1, CHECKS_HEADERS)

    trajectory_ws = wb["Trajectory"]
    snapshot_ws = wb["Snapshot"]
    quarters = data.quarters
    traj = trajectory.set_index(["ticker", "quarter"])
    base_rows = snapshot[snapshot["scenario"] == Scenario.BASE.value].set_index("ticker")
    bear_rows = snapshot[snapshot["scenario"] == Scenario.BEAR.value].set_index("ticker")
    bull_rows = snapshot[snapshot["scenario"] == Scenario.BULL.value].set_index("ticker")

    rows: list[tuple[str, str, float, Any, float, str]] = []
    for ticker_index, ticker in enumerate(TICKERS):
        block_top = 5 + TRAJECTORY_BLOCK_HEIGHT * ticker_index
        for quarter_index, quarter in enumerate(quarters):
            column = get_column_letter(3 + quarter_index)
            token = _quarter_token(quarter)
            for offset, suffix, label, number_format, tolerance in (
                (2, "ANNUALIZED-AI-CAPEX", "Annualized AI capex", NF_USD4, 1e-8),
                (3, "ANNUALIZED-AI-REVENUE", "Annualized AI revenue", NF_USD4, 1e-8),
                (4, "FORWARD-ROIC", "Forward ROIC", NF_PCT4, 1e-10),
                (5, "SPREAD", "Spread", NF_PCT4, 1e-10),
            ):
                field = TRAJECTORY_ROWS[offset][0]
                expected = float(traj.loc[(ticker, quarter), field])
                rendered = trajectory_ws[f"{column}{block_top + offset}"].value
                rows.append(
                    (
                        f"{ticker}-{token}-{suffix}",
                        f"{ticker} {quarter} {label}",
                        expected,
                        rendered,
                        tolerance,
                        number_format,
                    )
                )
        column = "BCDEF"[ticker_index]
        for key, suffix, label, field, source, number_format, tolerance in (
            ("revenue_proxy", "SNAPSHOT-ANNUALIZED-AI-REVENUE",
             "Snapshot annualized AI revenue", "ai_revenue_proxy_b", "base", NF_USD4, 1e-8),
            ("ai_capex", "SNAPSHOT-AI-CAPEX", "Snapshot AI capex",
             "ai_capex_b", "base", NF_USD4, 1e-8),
            ("roic_base", "SNAPSHOT-BASE-ROIC", "Snapshot base ROIC",
             "forward_roic", "base", NF_PCT4, 1e-10),
            ("spread_base", "SNAPSHOT-BASE-SPREAD", "Snapshot base spread",
             "spread", "base", NF_PCT4, 1e-10),
            ("spread_bear", "SNAPSHOT-BEAR-SPREAD", "Snapshot bear spread",
             "spread", "bear", NF_PCT4, 1e-10),
            ("spread_bull", "SNAPSHOT-BULL-SPREAD", "Snapshot bull spread",
             "spread", "bull", NF_PCT4, 1e-10),
            ("delta_yoy", "DELTA-SPREAD-YOY-BPS",
             "Year-over-year run-rate spread change in bps",
             "delta_spread_yoy_bps", "base", NF_DEC2, 0.01),
            ("delta_qoq", "DELTA-SPREAD-QOQ-BPS",
             "Sequential run-rate spread change in bps",
             "delta_spread_qoq_bps", "base", NF_DEC2, 0.01),
            ("delta_baseline", "DELTA-SPREAD-BPS",
             "Baseline-window run-rate spread change in bps",
             "delta_spread_bps", "base", NF_DEC2, 0.01),
        ):
            frame = {"base": base_rows, "bear": bear_rows, "bull": bull_rows}[source]
            expected = _clean(frame.loc[ticker, field])
            if expected is None:
                continue  # anchor does not exist yet; nothing to check
            rendered = snapshot_ws[f"{column}{SNAP_ROW[key]}"].value
            rows.append(
                (
                    f"{ticker}-{suffix}",
                    f"{ticker} {label}",
                    float(expected),
                    rendered,
                    tolerance,
                    number_format,
                )
            )

    failures: list[tuple[str, str]] = []
    for index, (check_id, label, expected, rendered, tolerance, number_format) in enumerate(rows):
        row = 4 + index
        ok = isinstance(rendered, (int, float)) and not isinstance(rendered, bool)
        ok = ok and abs(float(rendered) - expected) <= tolerance
        status = "PASS" if ok else "FAIL"
        if not ok:
            failures.append((check_id, f"expected {expected!r}, rendered {rendered!r}"))
        _put(ws, f"A{row}", check_id)
        _put(ws, f"B{row}", label)
        _put(ws, f"C{row}", expected, number_format=number_format)
        _put(
            ws,
            f"D{row}",
            float(rendered) if isinstance(rendered, (int, float)) else rendered,
            number_format=number_format,
        )
        _put(ws, f"E{row}", tolerance)
        _put(ws, f"F{row}", status, font=FONT_BODY if ok else FONT_BODY_BOLD)

    ws.freeze_panes = "A4"
    return failures


# --------------------------------------------------------------------------- #
# Assembly
# --------------------------------------------------------------------------- #


def output_path_for(data: DataLayer, out_dir: Path | str = DEFAULT_OUT_DIR) -> Path:
    """``build/ai_capex_forward_roic_analysis_v03_5q_through_Q2-26.xlsx``."""
    latest = data.quarters[-1].replace(" ", "-")
    name = (
        f"ai_capex_forward_roic_analysis_{GENERATOR_VERSION}"
        f"_{len(data.quarters)}q_through_{latest}.xlsx"
    )
    return Path(out_dir) / name


_MODIFIED_RE = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")


def _normalise_zip(path: Path) -> None:
    """Rewrite the .xlsx deterministically so identical inputs give identical bytes.

    Two things vary run to run and neither is content: openpyxl stamps every zip
    entry with the wall clock, and its writer overwrites ``dcterms:modified`` with
    ``utcnow()`` at save time regardless of what the properties say. Both are pinned
    here.
    """
    stamp = FIXED_TIMESTAMP.strftime("%Y-%m-%dT%H:%M:%SZ").encode("ascii")
    with zipfile.ZipFile(path, "r") as source:
        entries = [(info, source.read(info.filename)) for info in source.infolist()]
    temporary = path.with_suffix(".tmp")
    with zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as target:
        for info, payload in entries:
            if info.filename == "docProps/core.xml":
                payload = _MODIFIED_RE.sub(rb"\g<1>" + stamp + rb"\g<2>", payload)
            fixed = zipfile.ZipInfo(info.filename, date_time=ZIP_TIMESTAMP)
            fixed.compress_type = zipfile.ZIP_DEFLATED
            fixed.external_attr = info.external_attr
            fixed.create_system = 0
            target.writestr(fixed, payload)
    temporary.replace(path)


def build_workbook(
    data_dir: Path | str = DEFAULT_DATA_DIR,
    out_path: Path | str | None = None,
    out_dir: Path | str = DEFAULT_OUT_DIR,
    *,
    baseline_quarter: str | None = None,
    strict: bool = True,
) -> tuple[Path, dict[str, Any]]:
    """Render the workbook. Returns ``(path, summary)``."""
    data = DataLayer(data_dir)
    notes = NoteIndex(data)

    trajectory = mbuild.build_trajectory(data.facts, data.assumptions)
    latest_quarter = data.quarters[-1]
    baseline = baseline_quarter or data.quarters[0]
    snapshot = mbuild.build_snapshot(
        data.facts, data.assumptions, latest_quarter=latest_quarter, base_quarter=baseline
    )

    target = Path(out_path) if out_path else output_path_for(data, out_dir)
    target = target.resolve()
    if target == FROZEN_WORKBOOK.resolve():
        raise ValueError(
            "refusing to overwrite the frozen audit-of-record "
            f"{FROZEN_WORKBOOK.name}; choose another output path"
        )
    target.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheets = {name: wb.create_sheet(name) for name in SHEET_ORDER}
    for name, ws in sheets.items():
        _apply_sheet_chrome(ws, name)

    _build_trajectory(sheets["Trajectory"], data, notes, trajectory)
    _build_snapshot(
        sheets["Snapshot"], data, notes, snapshot, latest_quarter, baseline
    )
    _build_inputs(sheets["Inputs"], data, notes)
    _build_sources(sheets["Sources & Notes"], data, notes, len(data.quarters))
    failures = _build_checks(wb, sheets["Checks"], data, trajectory, snapshot)

    for key, value in DOC_PROPERTIES.items():
        setattr(wb.properties, key, value)
    wb.properties.created = FIXED_TIMESTAMP
    wb.properties.modified = FIXED_TIMESTAMP

    wb.save(target)
    _normalise_zip(target)

    note_count = sum(
        1
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.comment is not None
    )
    link_count = sum(
        1
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.hyperlink is not None
    )
    summary = {
        "path": target,
        "quarters": list(data.quarters),
        "latest_quarter": latest_quarter,
        "baseline_quarter": baseline,
        "notes": note_count,
        "hyperlinks": link_count,
        "checks": len(sheets["Checks"]["A"]) - 3,
        "check_failures": failures,
        "sha256": hashlib.sha256(target.read_bytes()).hexdigest(),
    }
    if failures and strict:
        raise AssertionError(
            f"{len(failures)} rendered value(s) disagree with the model: {failures[:5]}"
        )
    return target, summary


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--data-dir", default=str(DEFAULT_DATA_DIR))
    parser.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR))
    parser.add_argument("--out", default=None, help="explicit output path")
    parser.add_argument(
        "--baseline-quarter",
        default=None,
        help="permanent baseline for the fixed change row (default: oldest quarter)",
    )
    args = parser.parse_args(argv)

    frozen_before = (
        hashlib.sha256(FROZEN_WORKBOOK.read_bytes()).hexdigest()
        if FROZEN_WORKBOOK.exists()
        else None
    )
    path, summary = build_workbook(
        args.data_dir, args.out, args.out_dir, baseline_quarter=args.baseline_quarter
    )
    frozen_after = (
        hashlib.sha256(FROZEN_WORKBOOK.read_bytes()).hexdigest()
        if FROZEN_WORKBOOK.exists()
        else None
    )
    if frozen_before != frozen_after:
        raise AssertionError("the frozen v02 workbook changed during the build")

    print(f"wrote {path}")
    print(f"  sha256          {summary['sha256']}")
    print(f"  quarters ({len(summary['quarters'])})  {', '.join(summary['quarters'])}")
    print(f"  latest / base   {summary['latest_quarter']} / {summary['baseline_quarter']}")
    print(f"  cell notes      {summary['notes']}")
    print(f"  hyperlinks      {summary['hyperlinks']}")
    print(f"  checks          {summary['checks']} rows, {len(summary['check_failures'])} FAIL")
    if frozen_before:
        print(f"  v02 unchanged   {frozen_before}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
