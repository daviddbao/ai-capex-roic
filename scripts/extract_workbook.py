"""Extract ai_capex_forward_roic_analysis_v02.xlsx into a canonical CSV data layer.

READ-ONLY on the workbook: openpyxl load_workbook only, never wb.save().
The workbook is the audit-of-record; this script must never write to it.

Usage:  python scripts/extract_workbook.py
Writes: data/{facts,assumptions,sources,cell_notes,hyperlinks,provenance,
              formulas,expected_outputs}.csv
Column semantics are documented in docs/SCHEMA.md.
"""
import sys, csv, os, datetime
import openpyxl
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(REPO, "ai_capex_forward_roic_analysis_v02.xlsx")
DATA = os.path.join(REPO, "data")
DOCS = os.path.join(REPO, "docs")
os.makedirs(DATA, exist_ok=True)
os.makedirs(DOCS, exist_ok=True)

wbf = openpyxl.load_workbook(XLSX, data_only=False)   # formulas + comments + fills
wbv = openpyxl.load_workbook(XLSX, data_only=True)    # cached values

def w(path, header, rows):
    with open(os.path.join(DATA, path), 'w', newline='', encoding='utf-8') as fh:
        cw = csv.writer(fh, lineterminator='\n')
        cw.writerow(header)
        cw.writerows(rows)
    print(f"{path}: {len(rows)} data rows")
    return len(rows)

def num(v):
    """Full-precision, round-trippable rendering of a value."""
    if v is None:
        return ''
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, float):
        return repr(v)
    if isinstance(v, int):
        return str(v)
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d')
    return str(v)

counts = {}

# ---------------------------------------------------------------- 1. facts.csv
inp = wbv['Inputs']
FACT_COLS = [
    ('A', 'company'), ('B', 'ticker'), ('C', 'report_bucket'), ('D', 'fiscal_period'),
    ('E', 'period_end'), ('F', 'rpo_backlog_or_revenue_usd_b'), ('G', 'fact_metric'),
    ('H', 'quarterly_capex_usd_b'), ('I', 'capex_definition'), ('J', 'fact_source_url'),
    ('K', 'capex_source_url'), ('L', 'evidence_derivation'), ('M', 'fact_source_id'),
    ('N', 'capex_source_id'),
]
facts = []
for r in range(5, 30):
    facts.append([num(inp[f'{col}{r}'].value) for col, _ in FACT_COLS])
counts['facts.csv'] = w('facts.csv', [n for _, n in FACT_COLS], facts)

# earliest period covered, per ticker (for assumption versioning)
earliest = {}
for row in facts:
    t, pe = row[1], row[4]
    if t not in earliest or pe < earliest[t]:
        earliest[t] = pe

# ---------------------------------------------------- 2. assumptions.csv
ASSUM_COLS = [
    ('P', 'ticker'), ('Q', 'ai_revenue_proxy'), ('R', 'ai_share_of_rpo_revenue'),
    ('S', 'rpo_duration_years'), ('T', 'ai_share_of_capex'), ('U', 'nopat_margin_bear'),
    ('V', 'nopat_margin_base'), ('W', 'nopat_margin_bull'), ('X', 'wacc'),
    ('Y', 'damodaran_sector_date'),
    ('Z', 'annual_capex_guide_midpoint_actual_usd_b'), ('AA', 'plan_basis'),
    ('AB', 'plan_source_url'), ('AC', 'source_assumption_caveat'),
]
ticker_company = {row[1]: row[0] for row in facts}
arows = []
for r in range(5, 10):
    tk = inp[f'P{r}'].value
    vals = [num(inp[f'{col}{r}'].value) for col, _ in ASSUM_COLS]
    arows.append([tk, ticker_company[tk], earliest[tk], 'v02'] + vals[1:])
ahdr = ['ticker', 'company', 'effective_from', 'model_version'] + [n for _, n in ASSUM_COLS[1:]]
counts['assumptions.csv'] = w('assumptions.csv', ahdr, arows)

# ------------------------------------------------------- 3. sources.csv
src = wbv['Sources & Notes']
KIND = {'FACT': 'fact', 'CAPEX': 'capex', 'PLAN': 'plan', 'WACC': 'wacc'}
srows, seen = [], set()
r = 5
while src[f'A{r}'].value:
    sid = str(src[f'A{r}'].value).strip()
    assert sid not in seen, f'duplicate source_id {sid}'
    seen.add(sid)
    kind = KIND[sid.rsplit('-', 1)[1]]
    filing = num(src[f'G{r}'].value)
    metric = num(src[f'D{r}'].value)
    title = f'{filing} \u2014 {metric}' if filing else metric
    srows.append([
        sid, num(src[f'H{r}'].value), num(src[f'B{r}'].value), num(src[f'C{r}'].value),
        kind, title, num(src[f'J{r}'].value),
        # retained ledger detail (additive to the required 7 columns)
        num(src[f'E{r}'].value), num(src[f'F{r}'].value), num(src[f'I{r}'].value),
        num(src[f'K{r}'].value), num(src[f'L{r}'].value), 'yes',
    ])
    r += 1

# Two secondary sources are cited only inside cell notes / derivations and have no
# row in the workbook's own ledger. Recorded here with synthetic ids and flagged
# in_workbook_ledger=no so the registry is complete without pretending they were
# ledger entries.
srows.append([
    'ORCL-FY25Q3-10Q-DERIV',
    'https://www.sec.gov/Archives/edgar/data/1341439/000095017025037143/orcl-20250228.htm',
    'Oracle', 'FY25 Q3 / 2025-02-28', 'capex',
    'orcl-20250228.htm — nine-month cumulative capex used to derive ORCL Q2 25 quarter capex',
    '', '12.135', 'SEC filing',
    'Nine-month FY2025 cash capex of $12.135B; ORCL-Q225-CAPEX = $21.215B FY2025 less $12.135B.',
    'Cited in cell notes', 'Derivation input only; cited in the notes on Trajectory!C27, Inputs!H20, Inputs!K20, Sources & Notes!H42',
    'no'])
srows.append([
    'ORCL-Q4FY26-SLIDES',
    'https://s23.q4cdn.com/440135859/files/doc_financials/2026/q4/Q4-FY26-Oracle-Earnings-Slides.pdf',
    'Oracle', 'FY26 Q4 / 2026-05-31', 'plan',
    'Q4-FY26-Oracle-Earnings-Slides.pdf — official Q4 FY26 earnings slides (FY2027 net-cash-outlay capex guide)',
    '01_sources/company_filings/oracle/Oracle_2026-06-10_Q4-FY26_Earnings_Slides.pdf',
    '70', 'Official company disclosure',
    'Guides to approximately $70B of FY2027 net cash outlay for capex, a non-GAAP measure that is not interchangeable with gross capex; NOT used as the model denominator.',
    'Cited in cell notes',
    'Explicitly rejected as the snapshot denominator; cited in the notes on Snapshot!E9, Snapshot!E10, Inputs!Z8, Inputs!AB8, Sources & Notes!H51',
    'no'])

counts['sources.csv'] = w('sources.csv', [
    'source_id', 'url', 'company', 'period', 'kind', 'title_or_description',
    'local_path_if_any', 'reported_value', 'classification', 'evidence_derivation',
    'status', 'caveat', 'in_workbook_ledger'], srows)

# ---------------------------------------------------- 4a. cell_notes.csv
nrows = []
for ws in wbf.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if c.comment is not None:
                nrows.append([ws.title, c.coordinate, c.row, c.column_letter,
                              c.comment.author or '', c.comment.text])
counts['cell_notes.csv'] = w('cell_notes.csv',
    ['sheet', 'cell', 'row', 'column', 'author', 'note_text'], nrows)
url_notes = sum(1 for x in nrows if 'http' in x[5])
print(f'   ...of which URL-bearing: {url_notes}')

# ---------------------------------------------------- 4b. hyperlinks.csv
hrows = []
for ws in wbf.worksheets:
    wsv = wbv[ws.title]
    for row in ws.iter_rows():
        for c in row:
            if c.hyperlink is not None:
                h = c.hyperlink
                hrows.append([ws.title, c.coordinate, c.row, c.column_letter,
                              num(wsv[c.coordinate].value), h.display or '',
                              h.target or '', h.location or '', h.tooltip or ''])
counts['hyperlinks.csv'] = w('hyperlinks.csv',
    ['sheet', 'cell', 'row', 'column', 'cell_text', 'display', 'target',
     'location', 'tooltip'], hrows)

# ---------------------------------------------------- 5. provenance.csv
FILLCLASS = {'FFEAF2F8': 'fact', 'FFFFF2CC': 'assumption'}
FILLNAME = {'FFEAF2F8': 'light blue', 'FFFFF2CC': 'light yellow',
            'FF1F4E78': 'dark blue header', 'FFF2F2F2': 'light grey banner'}
inpf = wbf['Inputs']
prows = []
for r in range(1, inpf.max_row + 1):
    for ci in range(1, inpf.max_column + 1):
        cf = inpf.cell(row=r, column=ci)
        cv = inp.cell(row=r, column=ci)
        pt = cf.fill.patternType
        rgb = cf.fill.start_color.rgb if pt else None
        rgb = rgb if isinstance(rgb, str) else None
        if pt is None and cv.value is None:
            continue
        cls = FILLCLASS.get(rgb, 'other') if pt else 'other'
        # header/banner rows are chrome, never data provenance
        if r <= 4:
            cls = 'other'
        hdr = inp.cell(row=4, column=ci).value if r >= 5 else ''
        prows.append(['Inputs', cf.coordinate, r, get_column_letter(ci),
                      num(hdr), rgb or '', FILLNAME.get(rgb, '' if not rgb else 'other'),
                      cls, '1' if cv.value is not None else '0', num(cv.value)[:200]])
counts['provenance.csv'] = w('provenance.csv',
    ['sheet', 'cell', 'row', 'column', 'header', 'fill_rgb', 'fill_name',
     'fill_class', 'has_value', 'value_preview'], prows)
nf = sum(1 for x in prows if x[7] == 'fact')
na = sum(1 for x in prows if x[7] == 'assumption')
print(f'   fact-filled cells: {nf}  assumption-filled cells: {na}  other: {len(prows)-nf-na}')

# ---------------------------------------------------- 6. formulas.csv
frows = []
for name in ('Trajectory', 'Snapshot'):
    ws = wbf[name]
    wsv = wbv[name]
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                frows.append([name, c.coordinate, c.row, c.column_letter, c.value,
                              num(wsv[c.coordinate].value)])
counts['formulas.csv'] = w('formulas.csv',
    ['sheet', 'cell', 'row', 'column', 'formula', 'cached_value'], frows)

# ------------------------------------------- 7. expected_outputs.csv
traj = wbv['Trajectory']
QCOL = {'C': 'Q2 25', 'D': 'Q3 25', 'E': 'Q4 25', 'F': 'Q1 26', 'G': 'Q2 26'}
BLOCKS = [('MSFT', 5), ('GOOG', 12), ('AMZN', 19), ('ORCL', 26), ('META', 33)]
erows = []
for tk, top in BLOCKS:
    for off in range(6):
        r = top + off
        metric = traj[f'B{r}'].value
        scen = 'base' if off in (4, 5) else 'n/a'
        for col, per in QCOL.items():
            v = traj[f'{col}{r}'].value
            erows.append([tk, per, 'trajectory', scen, metric, num(v),
                          'number' if isinstance(v, (int, float)) else 'text'])

snap = wbv['Snapshot']
SNAP_SCEN = {15: 'base', 17: 'base', 19: 'base', 20: 'bear', 21: 'bear',
             22: 'bull', 23: 'bull'}
SNAP_PERIOD = {25: 'Q2 26', 26: 'Q2 25', 27: 'Q2 25 to Q2 26'}
SNAP_TCOL = {'B': 'MSFT', 'C': 'GOOG', 'D': 'AMZN', 'E': 'ORCL', 'F': 'META'}
for r in list(range(4, 24)) + [25, 26, 27]:
    metric = snap[f'A{r}'].value
    if metric is None or str(metric).startswith('\u2014'):
        continue
    scen = SNAP_SCEN.get(r, 'n/a')
    per = SNAP_PERIOD.get(r, 'Q2 26')
    for col, tk in SNAP_TCOL.items():
        v = snap[f'{col}{r}'].value
        erows.append([tk, per, 'snapshot', scen, metric, num(v),
                      'number' if isinstance(v, (int, float)) else 'text'])
counts['expected_outputs.csv'] = w('expected_outputs.csv',
    ['company', 'period', 'view', 'scenario', 'metric', 'value', 'value_type'], erows)

# ---------------------------------------------------------------- audit
print('\n--- audit counts ---')
tot_c = sum(1 for ws in wbf.worksheets for row in ws.iter_rows() for c in row if c.comment)
tot_h = sum(1 for ws in wbf.worksheets for row in ws.iter_rows() for c in row if c.hyperlink)
tot_f = sum(1 for ws in wbf.worksheets for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith('='))
print('comments (all sheets):', tot_c, ' URL-bearing:', url_notes)
print('hyperlinks (all sheets):', tot_h)
print('formulas (all sheets):', tot_f)
ck = wbv['Checks']
pas = sum(1 for row in ck.iter_rows() for c in row
          if isinstance(c.value, str) and c.value.strip() == 'PASS')
fail = sum(1 for row in ck.iter_rows() for c in row
           if isinstance(c.value, str) and c.value.strip() == 'FAIL')
print('checks PASS:', pas, 'FAIL:', fail)
for k, v in counts.items():
    print(f'  {k}: {v}')
