"""Acceptance tests for the values-only workbook generator.

The load-bearing test is :func:`test_expected_output_matches_v02`: every one of the
260 cached values in ``data/expected_outputs.csv`` -- the values Excel itself
computed in the frozen v02 workbook -- must come back out of the regenerated
workbook. A mismatch is a real finding about the renderer or the model, never a
reason to relax the expectation.

The second half of the file appends a synthetic sixth quarter to a *copy* of the
data layer and asserts the Trajectory widened by one column, nothing rolled off,
and the YoY/QoQ/baseline anchors rolled to the right quarters.
"""

from __future__ import annotations

import datetime as _dt
import hashlib
import shutil
import sys
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import pytest

REPO = Path(__file__).resolve().parents[2]
for candidate in (str(REPO), str(REPO / "scripts")):
    if candidate not in sys.path:
        sys.path.insert(0, candidate)

import build_workbook as bw  # noqa: E402
from model import build as mbuild  # noqa: E402

if hasattr(sys.stdout, "reconfigure"):  # the workbook carries x, delta, en-dashes
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

DATA_DIR = REPO / "data"
RTOL = 1e-12

# Trajectory metric label -> metric-row offset inside a company block. The first
# row's label differs per company, so it is resolved through the renderer's own map.
_TRAJECTORY_OFFSET_BY_LABEL = {
    "Quarterly Capex ($B)": 1,
    "Annualized AI Capex ($B)": 2,
    "Annualized AI Revenue Proxy ($B)": 3,
    "Forward ROIC (Base)": 4,
    "Spread vs WACC (ppt)": 5,
}

# v02 had one change row; the regenerated Snapshot has three. Where v02's single
# label is now ambiguous, name every generated row it corresponds to -- with five
# quarters on file the YoY anchor and the permanent baseline are the same quarter,
# so both rows must carry the v02 value.
_V02_SNAPSHOT_METRIC_CROSSWALK: dict[str, tuple[str, ...]] = {
    "Spread Q2 26 (run-rate basis, ppt)": ("rr_latest",),
    "Spread Q2 25 (run-rate basis, ppt)": ("rr_yoy", "rr_baseline"),
    "Δ Spread Q2 26 vs Q2 25 (bps)": ("delta_yoy", "delta_baseline"),
}


# --------------------------------------------------------------------------- #
# Fixtures
# --------------------------------------------------------------------------- #


@pytest.fixture(scope="session")
def frozen_hash_before() -> str:
    return hashlib.sha256(bw.FROZEN_WORKBOOK.read_bytes()).hexdigest()


@pytest.fixture(scope="session")
def generated(tmp_path_factory, frozen_hash_before: str) -> tuple[Path, dict[str, Any]]:
    """Build from the live 5-quarter data layer into a temporary directory."""
    out_dir = tmp_path_factory.mktemp("workbook")
    path, summary = bw.build_workbook(DATA_DIR, out_dir=out_dir)
    return path, summary


@pytest.fixture(scope="session")
def workbook(generated) -> openpyxl.Workbook:
    return openpyxl.load_workbook(generated[0])


@pytest.fixture(scope="session")
def expected_outputs() -> pd.DataFrame:
    # keep_default_na=False: META's duration is the literal string "N/A", which
    # pandas would otherwise silently turn into a missing value.
    return pd.read_csv(DATA_DIR / "expected_outputs.csv", keep_default_na=False)


@pytest.fixture(scope="session")
def data_layer() -> bw.DataLayer:
    return bw.DataLayer(DATA_DIR)


# --------------------------------------------------------------------------- #
# The frozen workbook must not move
# --------------------------------------------------------------------------- #


def test_frozen_workbook_is_byte_identical_after_a_build(
    generated, frozen_hash_before: str
) -> None:
    after = hashlib.sha256(bw.FROZEN_WORKBOOK.read_bytes()).hexdigest()
    assert after == frozen_hash_before
    assert after == "b0f2f3284fd37ac50653366af1e82a402c0a3bab96ddb619f64a6682d99ddd52"


def test_generator_refuses_to_write_over_the_frozen_workbook(tmp_path) -> None:
    with pytest.raises(ValueError, match="audit-of-record"):
        bw.build_workbook(DATA_DIR, out_path=bw.FROZEN_WORKBOOK)


def test_output_is_named_with_a_version_and_the_quarter_count(generated) -> None:
    name = generated[0].name
    assert name.startswith(f"ai_capex_forward_roic_analysis_{bw.GENERATOR_VERSION}_")
    assert "_5q_through_Q2-26" in name
    assert name != bw.FROZEN_WORKBOOK.name


# --------------------------------------------------------------------------- #
# Structure
# --------------------------------------------------------------------------- #


def test_sheet_names_and_order(workbook) -> None:
    assert workbook.sheetnames == list(bw.SHEET_ORDER)


def test_no_sheet_carries_a_model_formula(workbook) -> None:
    """Values only. The model lives in model/calc.py and nowhere else."""
    formulas = [
        f"{ws.title}!{cell.coordinate}"
        for ws in workbook.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    assert formulas == []


def test_trajectory_renders_one_column_per_quarter(workbook, data_layer) -> None:
    ws = workbook["Trajectory"]
    quarters = list(data_layer.quarters)
    headers = [ws.cell(row=4, column=3 + i).value for i in range(len(quarters))]
    assert headers == quarters
    assert ws.cell(row=4, column=3 + len(quarters)).value is None


def test_column_widths_and_freeze_panes_match_v02(workbook) -> None:
    assert workbook["Trajectory"].freeze_panes == "C4"
    assert workbook["Snapshot"].freeze_panes == "B4"
    assert workbook["Inputs"].freeze_panes == "A5"
    assert workbook["Sources & Notes"].freeze_panes == "A4"
    assert workbook["Checks"].freeze_panes == "A4"
    for sheet, widths in bw.COLUMN_WIDTHS.items():
        ws = workbook[sheet]
        for column, width in widths.items():
            assert ws.column_dimensions[column].width == width, f"{sheet}!{column}"


def test_number_formats_and_row_heights_match_v02(workbook) -> None:
    """Every cell v02 and the render share must display identically.

    Snapshot rows 26 and 27 are excluded: they held the baseline anchor and the one
    change row in v02 and now hold the YoY and QoQ anchors, so a bps format there
    would be wrong. Those two rows moved to 28 and 31, which are checked instead.
    """
    v02 = openpyxl.load_workbook(bw.FROZEN_WORKBOOK)
    try:
        problems = []
        shared_rows = {
            "Trajectory": range(1, 39),
            "Snapshot": [*range(1, 26), 28, 31],
            "Inputs": range(1, 30),
            "Sources & Notes": range(1, 75),
        }
        v02_row_for = {28: 26, 31: 27}
        for sheet, rows in shared_rows.items():
            old, new = v02[sheet], workbook[sheet]
            for row in rows:
                source_row = v02_row_for.get(row, row) if sheet == "Snapshot" else row
                if (
                    old.row_dimensions[source_row].height
                    != new.row_dimensions[row].height
                ):
                    problems.append(
                        f"{sheet} row {row} height "
                        f"{old.row_dimensions[source_row].height!r} != "
                        f"{new.row_dimensions[row].height!r}"
                    )
                for column in range(1, old.max_column + 1):
                    a = old.cell(row=source_row, column=column)
                    b = new.cell(row=row, column=column)
                    if a.value is None and b.value is None:
                        continue
                    if a.number_format != b.number_format:
                        problems.append(
                            f"{sheet}!{b.coordinate}: number format "
                            f"{a.number_format!r} != {b.number_format!r}"
                        )
                    if (a.font.b or False) != (b.font.b or False) or a.font.sz != b.font.sz:
                        problems.append(
                            f"{sheet}!{b.coordinate}: font "
                            f"{(a.font.sz, a.font.b)} != {(b.font.sz, b.font.b)}"
                        )
        assert not problems, "\n".join(problems[:20])
    finally:
        v02.close()


def test_the_three_change_rows_use_the_bps_format(workbook) -> None:
    ws = workbook["Snapshot"]
    for key in ("delta_yoy", "delta_qoq", "delta_baseline"):
        for column in "BCDEF":
            assert ws[f"{column}{bw.SNAP_ROW[key]}"].number_format == bw.NF_BPS
    for key in ("rr_latest", "rr_yoy", "rr_qoq", "rr_baseline"):
        for column in "BCDEF":
            assert ws[f"{column}{bw.SNAP_ROW[key]}"].number_format == bw.NF_PCT1


def test_banner_and_title_text_survive(workbook) -> None:
    assert workbook["Trajectory"]["A1"].value == bw.TRAJECTORY_TITLE
    assert workbook["Snapshot"]["A1"].value == bw.SNAPSHOT_TITLE
    assert workbook["Inputs"]["A1"].value == bw.INPUTS_TITLE_LEFT
    assert workbook["Inputs"]["P1"].value == bw.INPUTS_TITLE_RIGHT
    assert workbook["Sources & Notes"]["A1"].value == bw.SOURCES_TITLE
    assert workbook["Checks"]["A1"].value == bw.CHECKS_TITLE
    assert "analyst assumptions" in workbook["Inputs"]["A2"].value


# --------------------------------------------------------------------------- #
# Cell-value parity against the frozen workbook's cached values
# --------------------------------------------------------------------------- #


def _address_for_expected(row: pd.Series, quarters: list[str]) -> tuple[str, tuple[str, ...]]:
    """(sheet, addresses) the expected_outputs row should be found at."""
    ticker = str(row["company"])
    ticker_index = bw.TICKERS.index(ticker)
    if row["view"] == "trajectory":
        metric = str(row["metric"])
        if metric == bw.TRAJECTORY_FACT_LABEL[ticker]:
            offset = 0
        else:
            offset = _TRAJECTORY_OFFSET_BY_LABEL[metric]
        sheet_row = 5 + bw.TRAJECTORY_BLOCK_HEIGHT * ticker_index + offset
        column = openpyxl.utils.get_column_letter(3 + quarters.index(str(row["period"])))
        return "Trajectory", (f"{column}{sheet_row}",)

    column = "BCDEF"[ticker_index]
    metric = str(row["metric"])
    keys = _V02_SNAPSHOT_METRIC_CROSSWALK.get(metric)
    if keys is None:
        keys = (_SNAPSHOT_KEY_BY_V02_LABEL[metric],)
    return "Snapshot", tuple(f"{column}{bw.SNAP_ROW[key]}" for key in keys)


#: v02 Snapshot label -> the renderer's row key, for the rows v02 and the
#: regenerated sheet share verbatim.
_SNAPSHOT_KEY_BY_V02_LABEL = {
    "Latest Quarter": "latest_quarter",
    "AI Revenue Source": "revenue_source",
    "Total RPO/Backlog or Annualized Revenue ($B)": "fact",
    "AI Share % (assumption)": "ai_share",
    "AI-linked RPO/Revenue ($B)": "ai_linked",
    "Annual Capex Guide / Midpoint / Actual ($B)": "capex_guide",
    "Plan Basis": "plan_basis",
    "AI Share of Capex % (assumption)": "ai_share_capex",
    "AI Capex ($B)": "ai_capex",
    "RPO Duration (years, assumption)": "duration",
    "Annualized AI Revenue Proxy ($B)": "revenue_proxy",
    "NOPAT Margin (Base)": "nopat_base",
    "Forward ROIC (Base)": "roic_base",
    "WACC (Damodaran sector, Jan. 2026)": "wacc",
    "Spread (Base, ppt)": "spread_base",
    "Forward ROIC (Bear)": "roic_bear",
    "Spread (Bear, ppt)": "spread_bear",
    "Forward ROIC (Bull)": "roic_bull",
    "Spread (Bull, ppt)": "spread_bull",
}


def test_expected_output_row_count(expected_outputs) -> None:
    assert len(expected_outputs) == 260


def test_expected_output_matches_v02(workbook, expected_outputs, data_layer) -> None:
    """All 260 cached v02 values reproduce in the regenerated workbook."""
    quarters = list(data_layer.quarters)
    mismatches: list[str] = []
    compared = 0
    for _, row in expected_outputs.iterrows():
        sheet, addresses = _address_for_expected(row, quarters)
        ws = workbook[sheet]
        for address in addresses:
            actual = ws[address].value
            compared += 1
            if row["value_type"] == "number":
                expected = float(row["value"])
                if not isinstance(actual, (int, float)) or isinstance(actual, bool):
                    mismatches.append(
                        f"{sheet}!{address} ({row['company']} {row['metric']}): "
                        f"expected number {expected!r}, got {actual!r}"
                    )
                elif actual != pytest.approx(expected, rel=RTOL, abs=0.0):
                    mismatches.append(
                        f"{sheet}!{address} ({row['company']} {row['metric']}): "
                        f"v02 cached {expected!r}, rendered {actual!r} "
                        f"(abs diff {abs(float(actual) - expected)!r})"
                    )
            else:
                if str(actual) != str(row["value"]):
                    mismatches.append(
                        f"{sheet}!{address} ({row['company']} {row['metric']}): "
                        f"v02 cached {row['value']!r}, rendered {actual!r}"
                    )
    assert compared >= 260
    assert not mismatches, "\n".join(mismatches[:20])


def test_inputs_sheet_reproduces_v02_addresses(workbook) -> None:
    """With five quarters on file the Inputs grid is still v02's grid, cell for cell."""
    v02 = openpyxl.load_workbook(bw.FROZEN_WORKBOOK, data_only=True)
    try:
        old, new = v02["Inputs"], workbook["Inputs"]
        mismatches = []
        for row in old.iter_rows(min_row=1, max_row=29, max_col=29):
            for cell in row:
                if cell.value is None:
                    continue
                other = new[cell.coordinate].value
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    ok = isinstance(other, (int, float)) and other == pytest.approx(
                        cell.value, rel=RTOL, abs=0.0
                    )
                else:
                    ok = str(other) == str(cell.value)
                if not ok:
                    mismatches.append(f"Inputs!{cell.coordinate}: {cell.value!r} vs {other!r}")
        assert not mismatches, "\n".join(mismatches[:20])
    finally:
        v02.close()


def test_sources_ledger_reproduces_v02_addresses(workbook) -> None:
    v02 = openpyxl.load_workbook(bw.FROZEN_WORKBOOK, data_only=True)
    try:
        old, new = v02["Sources & Notes"], workbook["Sources & Notes"]
        mismatches = []
        for row in old.iter_rows(min_row=4, max_row=64, max_col=12):
            for cell in row:
                if cell.value is None:
                    continue
                other = new[cell.coordinate].value
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    ok = isinstance(other, (int, float)) and other == pytest.approx(
                        cell.value, rel=RTOL, abs=0.0
                    )
                else:
                    ok = str(other) == str(cell.value)
                if not ok:
                    mismatches.append(
                        f"Sources & Notes!{cell.coordinate}: {cell.value!r} vs {other!r}"
                    )
        assert not mismatches, "\n".join(mismatches[:20])
    finally:
        v02.close()


# --------------------------------------------------------------------------- #
# Audit trail: notes, hyperlinks, fills
# --------------------------------------------------------------------------- #


def _notes_of(workbook) -> dict[tuple[str, str], str]:
    return {
        (ws.title, cell.coordinate): cell.comment.text
        for ws in workbook.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.comment is not None
    }


def _links_of(workbook) -> dict[tuple[str, str], tuple[str, str | None]]:
    return {
        (ws.title, cell.coordinate): (cell.hyperlink.target, cell.hyperlink.display)
        for ws in workbook.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.hyperlink is not None
    }


def test_every_v02_cell_note_survives_at_its_address(workbook) -> None:
    """All 455 notes land on the cell they were attached to in v02."""
    stored = pd.read_csv(DATA_DIR / "cell_notes.csv")
    assert len(stored) == 455
    notes = _notes_of(workbook)
    missing = [
        (r.sheet, r.cell) for r in stored.itertuples() if (r.sheet, r.cell) not in notes
    ]
    assert not missing, f"{len(missing)} v02 notes lost: {missing[:10]}"


def test_v02_note_text_is_carried_verbatim_except_the_two_moved_change_rows(
    workbook,
) -> None:
    """445 of 455 notes are byte-identical to v02.

    The exceptions are Snapshot rows 26 and 27 (5 companies each). v02 had one
    change row; those two rows are now the YoY anchor and the QoQ anchor, and the
    baseline pair moved to rows 28 and 31. The notes that moved keep every URL and
    evidence sentence -- only the ``Formula:`` line changes, to echo the row's new,
    more specific label.
    """
    stored = pd.read_csv(DATA_DIR / "cell_notes.csv")
    notes = _notes_of(workbook)
    differing = [
        (r.sheet, r.cell)
        for r in stored.itertuples()
        if notes.get((r.sheet, r.cell)) != r.note_text
    ]
    assert set(differing) == {("Snapshot", f"{c}{r}") for c in "BCDEF" for r in (26, 27)}
    for sheet, cell in differing:
        original = stored[(stored.sheet == sheet) & (stored.cell == cell)].iloc[0].note_text
        rewritten = notes[(sheet, cell)]
        strip = lambda text: [  # noqa: E731
            line for line in text.split("\n") if not line.startswith("Formula: ")
        ]
        assert strip(rewritten) == strip(original), f"{sheet}!{cell} lost audit content"


def test_the_moved_change_rows_keep_their_notes(workbook) -> None:
    """The v02 baseline anchor / baseline delta notes are on rows 28 and 31 now."""
    stored = pd.read_csv(DATA_DIR / "cell_notes.csv")
    notes = _notes_of(workbook)
    for column in "BCDEF":
        for v02_row, new_row in ((26, 28), (27, 31)):
            original = stored[
                (stored.sheet == "Snapshot") & (stored.cell == f"{column}{v02_row}")
            ].iloc[0].note_text
            moved = notes[("Snapshot", f"{column}{new_row}")]
            keep = lambda text: [  # noqa: E731
                line for line in text.split("\n") if not line.startswith("Formula: ")
            ]
            assert keep(moved) == keep(original)


def test_note_count(workbook, generated) -> None:
    """455 carried forward plus 20 for the four new Snapshot rows x 5 companies."""
    notes = _notes_of(workbook)
    assert len(notes) == 475
    assert generated[1]["notes"] == 475
    by_sheet = pd.Series([sheet for sheet, _ in notes]).value_counts().to_dict()
    assert by_sheet == {
        "Trajectory": 150,
        "Inputs": 145,
        "Snapshot": 120,
        "Sources & Notes": 60,
    }


def test_every_note_is_url_bearing(workbook) -> None:
    notes = _notes_of(workbook)
    without = [key for key, text in notes.items() if "http" not in text]
    assert not without, f"notes with no public URL: {without[:10]}"


def test_hyperlinks_survive_exactly(workbook, generated) -> None:
    stored = pd.read_csv(DATA_DIR / "hyperlinks.csv")
    assert len(stored) == 240
    links = _links_of(workbook)
    assert len(links) == 240
    assert generated[1]["hyperlinks"] == 240
    problems = []
    for r in stored.itertuples():
        key = (r.sheet, r.cell)
        if key not in links:
            problems.append(f"{key} missing")
            continue
        target, display = links[key]
        expected_display = r.display if isinstance(r.display, str) else None
        if target != r.target:
            problems.append(f"{key} target {target!r} != {r.target!r}")
        if (display or None) != (expected_display or None):
            problems.append(f"{key} display {display!r} != {expected_display!r}")
    assert not problems, "\n".join(problems[:20])


def _fill_classes(workbook) -> dict[tuple[str, str], str]:
    classes = {bw.RGB_FACT: "fact", bw.RGB_ASSUMPTION: "assumption"}
    out = {}
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.fill.patternType is None:
                    continue
                rgb = cell.fill.start_color.rgb
                if isinstance(rgb, str) and rgb in classes:
                    out[(ws.title, cell.coordinate)] = classes[rgb]
    return out


def test_inputs_fill_language_matches_provenance(workbook) -> None:
    """Blue = sourced filing fact, yellow = analyst assumption, cell for cell."""
    provenance = pd.read_csv(DATA_DIR / "provenance.csv")
    expected = {
        r.cell: r.fill_class
        for r in provenance.itertuples()
        if r.fill_class in ("fact", "assumption")
    }
    assert sum(1 for v in expected.values() if v == "fact") == 115
    assert sum(1 for v in expected.values() if v == "assumption") == 30
    actual = {
        cell: klass
        for (sheet, cell), klass in _fill_classes(workbook).items()
        if sheet == "Inputs"
    }
    assert actual == expected


def test_presentation_sheet_fill_language_matches_v02(workbook) -> None:
    """The blue/yellow classes on Trajectory and Snapshot match v02's, cell for cell.

    Snapshot rows 25+ carry no fill in either workbook, so widening that block does
    not disturb the comparison.
    """
    v02 = openpyxl.load_workbook(bw.FROZEN_WORKBOOK)
    try:
        old = _fill_classes(v02)
        new = _fill_classes(workbook)
        for sheet in ("Trajectory", "Snapshot"):
            old_sheet = {k: v for k, v in old.items() if k[0] == sheet}
            new_sheet = {k: v for k, v in new.items() if k[0] == sheet}
            assert new_sheet == old_sheet, sheet
    finally:
        v02.close()


def test_total_fill_counts(workbook) -> None:
    classes = _fill_classes(workbook)
    assert sum(1 for v in classes.values() if v == "fact") == 300
    assert sum(1 for v in classes.values() if v == "assumption") == 52


# --------------------------------------------------------------------------- #
# Checks sheet
# --------------------------------------------------------------------------- #


def test_checks_sheet_all_pass(workbook, generated) -> None:
    ws = workbook["Checks"]
    statuses = [
        ws[f"F{row}"].value for row in range(4, ws.max_row + 1) if ws[f"A{row}"].value
    ]
    assert statuses, "no checks rendered"
    assert set(statuses) == {"PASS"}
    assert generated[1]["check_failures"] == []


def test_checks_sheet_covers_every_quarter_and_the_new_anchors(workbook, data_layer) -> None:
    ws = workbook["Checks"]
    ids = [ws[f"A{row}"].value for row in range(4, ws.max_row + 1) if ws[f"A{row}"].value]
    # 4 trajectory checks per company-quarter + 9 snapshot checks per company
    assert len(ids) == 5 * (4 * len(data_layer.quarters) + 9) == 145
    assert "MSFT-Q226-SPREAD" in ids
    assert "MSFT-DELTA-SPREAD-BPS" in ids
    assert "MSFT-DELTA-SPREAD-YOY-BPS" in ids
    assert "MSFT-DELTA-SPREAD-QOQ-BPS" in ids


# --------------------------------------------------------------------------- #
# Snapshot anchors
# --------------------------------------------------------------------------- #


def test_snapshot_change_rows_are_labelled_unambiguously(workbook) -> None:
    ws = workbook["Snapshot"]
    labels = {key: ws[f"A{row}"].value for key, row in bw.SNAP_ROW.items()}
    assert labels["rr_latest"] == "Spread Q2 26 (run-rate basis, ppt)"
    assert labels["rr_yoy"] == "Spread Q2 25 (run-rate basis, ppt) — YoY anchor"
    assert labels["rr_qoq"] == "Spread Q1 26 (run-rate basis, ppt) — QoQ anchor"
    assert labels["rr_baseline"] == "Spread Q2 25 (run-rate basis, ppt) — baseline anchor"
    assert labels["delta_yoy"] == "Δ Spread Q2 26 vs Q2 25 (YoY, bps)"
    assert labels["delta_qoq"] == "Δ Spread Q2 26 vs Q1 26 (QoQ, bps)"
    assert labels["delta_baseline"] == "Δ Spread Q2 26 vs Q2 25 (baseline, bps)"
    # every label on the sheet is distinct, so no row can be mistaken for another
    all_labels = [ws[f"A{row}"].value for row in range(4, 32)]
    assert len(set(all_labels)) == len(all_labels)


@pytest.mark.parametrize(
    "ticker,published_ppt",
    # ai_capex_forward_roic_analysis_v02_methodology.md, "Sequential spread change".
    # The doc published these; the workbook never computed them.
    [("MSFT", -5.5), ("GOOG", -3.8), ("AMZN", 1.2), ("ORCL", 7.7), ("META", -6.3)],
)
def test_qoq_row_reconciles_to_the_published_methodology_figure(
    workbook, ticker: str, published_ppt: float
) -> None:
    column = "BCDEF"[bw.TICKERS.index(ticker)]
    rendered_bps = workbook["Snapshot"][f"{column}{bw.SNAP_ROW['delta_qoq']}"].value
    assert rendered_bps / 100 == pytest.approx(published_ppt, abs=0.05)


def test_snapshot_qoq_row_matches_the_trajectory_sheet(workbook) -> None:
    """The QoQ row is the sequential change the methodology doc publishes."""
    trajectory = workbook["Trajectory"]
    snapshot = workbook["Snapshot"]
    for index, ticker in enumerate(bw.TICKERS):
        block = 5 + bw.TRAJECTORY_BLOCK_HEIGHT * index
        latest = trajectory[f"G{block + 5}"].value  # Q2 26 spread
        prior = trajectory[f"F{block + 5}"].value  # Q1 26 spread
        column = "BCDEF"[index]
        rendered = snapshot[f"{column}{bw.SNAP_ROW['delta_qoq']}"].value
        assert rendered == pytest.approx((latest - prior) * 10000, rel=RTOL, abs=0.0)


# --------------------------------------------------------------------------- #
# Determinism
# --------------------------------------------------------------------------- #


def test_two_builds_are_byte_identical(tmp_path) -> None:
    first, _ = bw.build_workbook(DATA_DIR, out_path=tmp_path / "a.xlsx")
    second, _ = bw.build_workbook(DATA_DIR, out_path=tmp_path / "b.xlsx")
    assert hashlib.sha256(first.read_bytes()).hexdigest() == (
        hashlib.sha256(second.read_bytes()).hexdigest()
    )


# --------------------------------------------------------------------------- #
# Growth: a sixth quarter, appended to a COPY of the data layer
# --------------------------------------------------------------------------- #


def _add_months(date: _dt.date, months: int) -> _dt.date:
    """Month arithmetic that keeps an end-of-month date at end of month."""
    month = date.month - 1 + months
    year = date.year + month // 12
    month = month % 12 + 1
    last_day = (
        _dt.date(year + (month == 12), month % 12 + 1, 1) - _dt.timedelta(days=1)
    ).day
    return _dt.date(year, month, min(date.day, last_day))


def _append_synthetic_quarter(source: Path, destination: Path, bucket: str = "Q3 26") -> Path:
    """Copy ``data/`` and append one more quarter per company. Never touches source."""
    shutil.copytree(source, destination)
    facts = pd.read_csv(destination / "facts.csv")
    sources = pd.read_csv(destination / "sources.csv")

    latest_bucket = mbuild.ordered_quarters(
        mbuild.load_inputs_from_csv(source)[0]
    )[-1]
    token = bucket.replace(" ", "")

    new_facts, new_sources = [], []
    for ticker in bw.TICKERS:
        row = facts[(facts.ticker == ticker) & (facts.report_bucket == latest_bucket)]
        row = row.iloc[0].to_dict()
        period_end = _add_months(_dt.date.fromisoformat(str(row["period_end"])), 3)
        fact_id, capex_id = f"{ticker}-{token}-FACT", f"{ticker}-{token}-CAPEX"
        fact_value = round(float(row["rpo_backlog_or_revenue_usd_b"]) * 1.10, 3)
        capex_value = round(float(row["quarterly_capex_usd_b"]) * 1.05, 3)
        row.update(
            {
                "report_bucket": bucket,
                "fiscal_period": f"{bucket} (synthetic)",
                "period_end": period_end.isoformat(),
                "rpo_backlog_or_revenue_usd_b": fact_value,
                "quarterly_capex_usd_b": capex_value,
                "fact_source_id": fact_id,
                "capex_source_id": capex_id,
            }
        )
        new_facts.append(row)
        for source_id, url, metric, value in (
            (fact_id, row["fact_source_url"], row["fact_metric"], fact_value),
            (capex_id, row["capex_source_url"], "Quarterly capex", capex_value),
        ):
            new_sources.append(
                {
                    "source_id": source_id,
                    "url": url,
                    "company": row["company"],
                    "period": f"{bucket} / {period_end.isoformat()}",
                    "kind": "fact" if source_id.endswith("FACT") else "capex",
                    "title_or_description": f"synthetic-{token}.htm — {metric}",
                    "local_path_if_any": "",
                    "reported_value": value,
                    "classification": "Synthetic test fixture",
                    "evidence_derivation": f"Synthetic {bucket} value for test purposes.",
                    "status": "Test fixture",
                    "caveat": "Not a real filing; test fixture only.",
                    "in_workbook_ledger": "yes",
                }
            )

    pd.concat([facts, pd.DataFrame(new_facts)], ignore_index=True).to_csv(
        destination / "facts.csv", index=False, lineterminator="\n"
    )
    pd.concat([sources, pd.DataFrame(new_sources)], ignore_index=True).to_csv(
        destination / "sources.csv", index=False, lineterminator="\n"
    )
    return destination


@pytest.fixture(scope="session")
def six_quarters(tmp_path_factory) -> tuple[Path, openpyxl.Workbook, dict[str, Any]]:
    root = tmp_path_factory.mktemp("six")
    data_dir = _append_synthetic_quarter(DATA_DIR, root / "data")
    path, summary = bw.build_workbook(data_dir, out_dir=root / "build")
    return path, openpyxl.load_workbook(path), summary


def test_appending_a_quarter_leaves_the_real_data_layer_alone(six_quarters) -> None:
    facts = pd.read_csv(DATA_DIR / "facts.csv")
    assert len(facts) == 25
    assert set(facts.report_bucket) == set(mbuild.QUARTERS)


def test_trajectory_widens_by_one_column_and_drops_nothing(six_quarters) -> None:
    _, workbook, _ = six_quarters
    ws = workbook["Trajectory"]
    headers = [ws.cell(row=4, column=3 + i).value for i in range(6)]
    assert headers == ["Q2 25", "Q3 25", "Q4 25", "Q1 26", "Q2 26", "Q3 26"]
    assert ws.cell(row=4, column=9).value is None
    # every original quarter's numbers are untouched in place
    original = openpyxl.load_workbook(bw.FROZEN_WORKBOOK, data_only=True)["Trajectory"]
    for index in range(len(bw.TICKERS)):
        block = 5 + bw.TRAJECTORY_BLOCK_HEIGHT * index
        for offset in range(6):
            for column in "CDEFG":
                assert ws[f"{column}{block + offset}"].value == pytest.approx(
                    original[f"{column}{block + offset}"].value, rel=RTOL, abs=0.0
                ), f"{column}{block + offset}"


def test_inputs_grows_by_one_row_per_company(six_quarters) -> None:
    _, workbook, _ = six_quarters
    ws = workbook["Inputs"]
    buckets = [ws[f"C{row}"].value for row in range(5, 35)]
    assert len(buckets) == 30
    assert buckets[:6] == ["Q2 25", "Q3 25", "Q4 25", "Q1 26", "Q2 26", "Q3 26"]
    assert ws["B11"].value == "GOOG"  # blocks shifted down by one row each
    assert ws["C35"].value is None


def test_output_filename_reflects_the_new_quarter(six_quarters) -> None:
    path, _, summary = six_quarters
    assert "_6q_through_Q3-26" in path.name
    assert summary["quarters"][-1] == "Q3 26"


def test_yoy_and_qoq_anchors_roll_forward(six_quarters) -> None:
    _, workbook, _ = six_quarters
    ws = workbook["Snapshot"]
    labels = {key: ws[f"A{row}"].value for key, row in bw.SNAP_ROW.items()}
    assert labels["rr_latest"] == "Spread Q3 26 (run-rate basis, ppt)"
    assert labels["rr_yoy"] == "Spread Q3 25 (run-rate basis, ppt) — YoY anchor"
    assert labels["rr_qoq"] == "Spread Q2 26 (run-rate basis, ppt) — QoQ anchor"
    # the baseline stays put; that is the whole point of it
    assert labels["rr_baseline"] == "Spread Q2 25 (run-rate basis, ppt) — baseline anchor"
    assert labels["delta_yoy"] == "Δ Spread Q3 26 vs Q3 25 (YoY, bps)"
    assert labels["delta_qoq"] == "Δ Spread Q3 26 vs Q2 26 (QoQ, bps)"
    assert labels["delta_baseline"] == "Δ Spread Q3 26 vs Q2 25 (baseline, bps)"


def test_rolled_anchors_carry_the_right_numbers(six_quarters) -> None:
    _, workbook, _ = six_quarters
    trajectory = workbook["Trajectory"]
    snapshot = workbook["Snapshot"]
    quarter_column = dict(zip(("Q2 25", "Q3 25", "Q4 25", "Q1 26", "Q2 26", "Q3 26"), "CDEFGH"))
    for index, ticker in enumerate(bw.TICKERS):
        spread_row = 5 + bw.TRAJECTORY_BLOCK_HEIGHT * index + 5
        column = "BCDEF"[index]

        def spread(quarter: str) -> float:
            return trajectory[f"{quarter_column[quarter]}{spread_row}"].value

        assert snapshot[f"{column}{bw.SNAP_ROW['rr_latest']}"].value == pytest.approx(
            spread("Q3 26"), rel=RTOL, abs=0.0
        )
        assert snapshot[f"{column}{bw.SNAP_ROW['rr_yoy']}"].value == pytest.approx(
            spread("Q3 25"), rel=RTOL, abs=0.0
        )
        assert snapshot[f"{column}{bw.SNAP_ROW['rr_qoq']}"].value == pytest.approx(
            spread("Q2 26"), rel=RTOL, abs=0.0
        )
        for key, anchor in (
            ("delta_yoy", "Q3 25"),
            ("delta_qoq", "Q2 26"),
            ("delta_baseline", "Q2 25"),
        ):
            assert snapshot[f"{column}{bw.SNAP_ROW[key]}"].value == pytest.approx(
                (spread("Q3 26") - spread(anchor)) * 10000, rel=RTOL, abs=0.0
            )


def test_quarter_local_notes_survive_the_widening(six_quarters) -> None:
    """Notes that describe a specific quarter are untouched by appending another.

    The Inputs assumption notes and the Snapshot notes legitimately re-point at the
    new latest quarter, so they are excluded; everything else must be verbatim.
    """
    _, workbook, _ = six_quarters
    stored = pd.read_csv(DATA_DIR / "cell_notes.csv")
    notes = _notes_of(workbook)
    quarter_local = stored[
        (stored.sheet == "Trajectory")
        | ((stored.sheet == "Inputs") & stored.column.isin(["F", "H", "J", "K"]))
    ]
    assert len(quarter_local) == 250
    survivors = {text for text in notes.values()}
    lost = [
        (r.sheet, r.cell) for r in quarter_local.itertuples() if r.note_text not in survivors
    ]
    assert not lost, f"{len(lost)} quarter-local notes lost: {lost[:10]}"


def test_new_quarter_cells_get_notes_and_links(six_quarters) -> None:
    _, workbook, _ = six_quarters
    trajectory = workbook["Trajectory"]
    for index in range(len(bw.TICKERS)):
        block = 5 + bw.TRAJECTORY_BLOCK_HEIGHT * index
        for offset in range(6):
            cell = trajectory[f"H{block + offset}"]
            assert cell.comment is not None, f"H{block + offset} has no note"
            assert "http" in cell.comment.text
            assert "Q3 26" in cell.comment.text
            if offset in (0, 1):
                assert cell.hyperlink is not None


def test_checks_grow_with_the_series(six_quarters) -> None:
    _, workbook, summary = six_quarters
    ws = workbook["Checks"]
    ids = [ws[f"A{row}"].value for row in range(4, ws.max_row + 1) if ws[f"A{row}"].value]
    assert len(ids) == 5 * (4 * 6 + 9) == 165
    assert "MSFT-Q326-SPREAD" in ids
    statuses = [ws[f"F{row}"].value for row in range(4, 4 + len(ids))]
    assert set(statuses) == {"PASS"}
    assert summary["check_failures"] == []


def test_caveat_bullet_reports_the_real_quarter_count(six_quarters, workbook) -> None:
    _, six, _ = six_quarters
    def bullet(wb):
        ws = wb["Sources & Notes"]
        return [
            ws[f"A{row}"].value
            for row in range(1, ws.max_row + 1)
            if isinstance(ws[f"A{row}"].value, str)
            and ws[f"A{row}"].value.startswith("• ")
            and "quarters are shown" in ws[f"A{row}"].value
        ][0]
    assert bullet(workbook).startswith("• 5 quarters are shown")
    assert bullet(six).startswith("• 6 quarters are shown")


# --------------------------------------------------------------------------- #
# Short history: anchors that do not exist yet
# --------------------------------------------------------------------------- #


@pytest.fixture(scope="session")
def two_quarters(tmp_path_factory) -> openpyxl.Workbook:
    root = tmp_path_factory.mktemp("short")
    data_dir = root / "data"
    shutil.copytree(DATA_DIR, data_dir)
    facts = pd.read_csv(data_dir / "facts.csv")
    facts = facts[facts.report_bucket.isin(["Q2 25", "Q3 25"])]
    facts.to_csv(data_dir / "facts.csv", index=False, lineterminator="\n")
    path, _ = bw.build_workbook(data_dir, out_dir=root / "build")
    return openpyxl.load_workbook(path)


def test_missing_yoy_anchor_renders_as_na_not_a_number(two_quarters) -> None:
    ws = two_quarters["Snapshot"]
    assert ws[f"A{bw.SNAP_ROW['rr_yoy']}"].value == "Spread (YoY anchor) — not yet available"
    assert ws[f"A{bw.SNAP_ROW['delta_yoy']}"].value == "Δ Spread (YoY, bps) — not yet available"
    for column in "BCDEF":
        assert ws[f"{column}{bw.SNAP_ROW['rr_yoy']}"].value == "n/a"
        assert ws[f"{column}{bw.SNAP_ROW['delta_yoy']}"].value == "n/a"
        # the sequential anchor does exist with two quarters on file
        assert isinstance(ws[f"{column}{bw.SNAP_ROW['delta_qoq']}"].value, float)


def test_short_history_skips_the_yoy_check_rather_than_inventing_one(two_quarters) -> None:
    ws = two_quarters["Checks"]
    ids = [ws[f"A{row}"].value for row in range(4, ws.max_row + 1) if ws[f"A{row}"].value]
    assert "MSFT-DELTA-SPREAD-YOY-BPS" not in ids
    assert "MSFT-DELTA-SPREAD-QOQ-BPS" in ids
    assert set(ws[f"F{row}"].value for row in range(4, 4 + len(ids))) == {"PASS"}
