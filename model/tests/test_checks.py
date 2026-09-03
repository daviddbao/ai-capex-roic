"""Port of the workbook's ``Checks`` sheet to pytest.

Structure of that sheet (header on row 3, checks on rows 4-138):

    A  Check ID          e.g. "MSFT-Q226-FORWARD-ROIC"
    B  Test              human-readable description
    C  Expected          a HARD-CODED literal, independently recalculated by
                         the workbook's author -- not a formula
    D  Workbook Value    a formula of the form ``=Trajectory!G9`` or
                         ``=Snapshot!E17`` pointing at one presentation cell
    E  Tolerance         1e-8 for $B amounts, 1e-10 for decimal ratios,
                         0.01 for the basis-point rows
    F  Status            ``=IF(ABS(C-D)<=E,"PASS","FAIL")``

The port keeps column C (the independent expected value) and column E (the
author's stated tolerance) but replaces column D with this implementation's
recomputation of the cell that D references. So each check asks the same
question the workbook asked, of Python instead of Excel.
"""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

import openpyxl
import pytest

from model import build

if hasattr(sys.stdout, "reconfigure"):  # Windows console is cp1252
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

WORKBOOK: Path = build.WORKBOOK_PATH

#: Rows 4..138 inclusive on the Checks sheet.
_FIRST_CHECK_ROW = 4
_EXPECTED_CHECK_COUNT = 135

#: ``=Trajectory!G9`` / ``=Snapshot!$E$17`` / ``='Sheet Name'!A1``
_REF_RE = re.compile(r"^=\s*'?(?P<sheet>[^'!]+)'?!\$?(?P<col>[A-Z]{1,3})\$?(?P<row>\d+)$")


@dataclass(frozen=True)
class Check:
    """One row of the Checks sheet."""

    row: int
    check_id: str
    description: str
    expected: float
    address: str
    tolerance: float
    workbook_status: str

    def __str__(self) -> str:  # keeps pytest ids readable
        return self.check_id


def _parse_reference(formula: object, row: int) -> str:
    if not isinstance(formula, str):
        raise AssertionError(f"Checks!D{row} is not a formula: {formula!r}")
    match = _REF_RE.match(formula.strip())
    if match is None:
        raise AssertionError(f"Checks!D{row} is not a simple cell reference: {formula!r}")
    return f"{match['sheet']}!{match['col']}{match['row']}"


@lru_cache(maxsize=1)
def load_checks() -> tuple[Check, ...]:
    """Read the Checks sheet: formulas for column D, cached values elsewhere."""
    formulas = openpyxl.load_workbook(WORKBOOK, data_only=False)
    values = openpyxl.load_workbook(WORKBOOK, data_only=True)
    try:
        fs = formulas["Checks"]
        vs = values["Checks"]
        checks: list[Check] = []
        for row in range(_FIRST_CHECK_ROW, fs.max_row + 1):
            check_id = vs.cell(row, 1).value
            if check_id is None:
                continue
            checks.append(
                Check(
                    row=row,
                    check_id=str(check_id),
                    description=str(vs.cell(row, 2).value),
                    expected=float(vs.cell(row, 3).value),
                    address=_parse_reference(fs.cell(row, 4).value, row),
                    tolerance=float(vs.cell(row, 5).value),
                    workbook_status=str(vs.cell(row, 6).value),
                )
            )
        return tuple(checks)
    finally:
        formulas.close()
        values.close()


@lru_cache(maxsize=1)
def recomputed_cells() -> dict[str, float]:
    """This implementation's values, keyed by workbook cell address."""
    trajectory, snapshot = build.build_all(WORKBOOK)
    return build.cell_map(trajectory, snapshot)


def test_checks_sheet_has_the_documented_number_of_checks() -> None:
    """The methodology doc claims 135 independent numerical checks."""
    checks = load_checks()
    assert len(checks) == _EXPECTED_CHECK_COUNT, (
        f"expected {_EXPECTED_CHECK_COUNT} checks, found {len(checks)}"
    )
    assert len({c.check_id for c in checks}) == len(checks), "duplicate check IDs"


def test_all_checks_reference_a_cell_this_model_computes() -> None:
    """No check may point at a cell outside the reimplemented surface."""
    ours = recomputed_cells()
    unmapped = sorted({c.address for c in load_checks() if c.address not in ours})
    assert not unmapped, f"checks reference cells with no recomputation: {unmapped}"


def test_workbook_reported_all_checks_passing() -> None:
    """Sanity: the cached Status column should be all PASS before we port it."""
    failures = [c.check_id for c in load_checks() if c.workbook_status != "PASS"]
    assert not failures, f"workbook's own Status column reports FAIL for: {failures}"


@pytest.mark.parametrize("check", load_checks(), ids=lambda c: c.check_id)
def test_check(check: Check) -> None:
    """Recompute the referenced cell and apply the sheet's own tolerance."""
    ours = recomputed_cells()
    assert check.address in ours, f"{check.check_id}: no value for {check.address}"
    actual = ours[check.address]
    delta = abs(check.expected - actual)
    assert delta <= check.tolerance, (
        f"{check.check_id} ({check.description}) at {check.address}: "
        f"expected {check.expected!r}, recomputed {actual!r}, "
        f"|delta| {delta!r} > tolerance {check.tolerance!r}"
    )


@pytest.mark.parametrize("check", load_checks(), ids=lambda c: c.check_id)
def test_check_is_also_tight(check: Check) -> None:
    """The sheet's tolerances are loose; our arithmetic should be far tighter.

    Column C is stored to 16 significant digits, so the only permissible
    difference is that last-digit rounding: rtol 1e-15.
    """
    actual = recomputed_cells()[check.address]
    assert actual == pytest.approx(check.expected, rel=1e-15, abs=1e-12)
